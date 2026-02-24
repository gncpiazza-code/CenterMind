# -*- coding: utf-8 -*-
"""
CenterMind — Reportes
=====================
Filtros: fecha, vendedor, estado, tipo PDV, número de cliente.
Vista previa en tabla + gráficos + exportar a Excel.
"""

from __future__ import annotations

import io
import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

import streamlit as st

# ─── Guard de sesión ──────────────────────────────────────────────────────────
if "logged_in" not in st.session_state or not st.session_state.logged_in:
    st.switch_page("app.py")

# ─── Configuración de página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="CenterMind · Reportes",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
DB_PATH  = BASE_DIR.parent.parent / "base_datos" / "centermind.db"

# ─── Imports opcionales ───────────────────────────────────────────────────────
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── CSS ──────────────────────────────────────────────────────────────────────
STYLE = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');

*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

html, body, [data-testid="stAppViewContainer"], [data-testid="stMain"] {
    background: #07080f !important;
    color: #e2e8f0 !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-testid="stHeader"],
[data-testid="stToolbar"],
[data-testid="stDecoration"]        { display: none !important; }
[data-testid="stMainBlockContainer"]{ padding: 0 !important; max-width: 100% !important; }
section[data-testid="stSidebar"]    { display: none !important; }
.block-container { padding: 0 !important; max-width: 100% !important; }

[data-testid="stAppViewContainer"]::before {
    content: '';
    position: fixed; inset: 0; z-index: 0;
    background:
        radial-gradient(ellipse 80% 50% at 10% 20%, rgba(251,191,36,0.03) 0%, transparent 60%),
        radial-gradient(ellipse 60% 40% at 90% 80%, rgba(34,211,238,0.02) 0%, transparent 60%),
        repeating-linear-gradient(0deg, transparent, transparent 39px, rgba(255,255,255,0.012) 40px),
        repeating-linear-gradient(90deg, transparent, transparent 39px, rgba(255,255,255,0.012) 40px);
    pointer-events: none;
}

/* ── Topbar ───────────────────────────────────────────────── */
.topbar {
    display: flex; align-items: center; justify-content: space-between;
    padding: 14px 32px;
    background: rgba(10,12,22,0.9);
    border-bottom: 1px solid rgba(251,191,36,0.12);
    backdrop-filter: blur(8px);
    position: sticky; top: 0; z-index: 100;
    margin-bottom: 0;
}
.topbar-logo {
    font-family: 'Bebas Neue', sans-serif;
    font-size: 26px; letter-spacing: 3px; color: #fbbf24;
}
.topbar-meta { font-size: 12px; color: rgba(226,232,240,0.4); letter-spacing: 1px; }
.user-badge {
    display: inline-flex; align-items: center; gap: 6px;
    padding: 4px 14px; border-radius: 999px;
    font-size: 11px; letter-spacing: 1px;
    background: rgba(34,211,238,0.1);
    border: 1px solid rgba(34,211,238,0.25);
    color: #22d3ee;
}

/* ── Cards ────────────────────────────────────────────────── */
.card {
    background: rgba(18,27,46,0.85);
    border: 1px solid rgba(34,48,76,0.9);
    border-radius: 16px;
    padding: 20px 22px;
    margin-bottom: 16px;
}
.card-title {
    font-size: 10px; letter-spacing: 3px; text-transform: uppercase;
    color: rgba(226,232,240,0.35); margin-bottom: 16px;
    display: flex; align-items: center; gap: 8px;
}
.card-title::after {
    content: ''; flex: 1; height: 1px;
    background: rgba(255,255,255,0.06);
}

/* ── KPI chips ────────────────────────────────────────────── */
.kpi-row {
    display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 4px;
}
.kpi-chip {
    display: flex; flex-direction: column; align-items: center;
    padding: 14px 20px; border-radius: 12px;
    border: 1px solid rgba(255,255,255,0.06);
    background: rgba(15,23,42,0.7);
    min-width: 100px;
}
.kpi-chip-num {
    font-family: 'Bebas Neue', sans-serif; font-size: 36px; line-height: 1;
}
.kpi-chip-lbl {
    font-size: 9px; letter-spacing: 2px; text-transform: uppercase;
    color: rgba(148,163,184,0.7); margin-top: 2px;
}
.kpi-green  { border-color: rgba(34,197,94,0.2);  }
.kpi-green  .kpi-chip-num { color: #22c55e; }
.kpi-amber  { border-color: rgba(252,211,77,0.2);  }
.kpi-amber  .kpi-chip-num { color: #fcd34d; }
.kpi-red    { border-color: rgba(239,68,68,0.2);   }
.kpi-red    .kpi-chip-num { color: #ef4444; }
.kpi-cyan   { border-color: rgba(34,211,238,0.2);  }
.kpi-cyan   .kpi-chip-num { color: #22d3ee; }
.kpi-muted  { border-color: rgba(148,163,184,0.15);}
.kpi-muted  .kpi-chip-num { color: #94a3b8; }

/* ── Tabla ────────────────────────────────────────────────── */
div[data-testid="stDataFrame"] {
    border-radius: 12px !important;
    overflow: hidden !important;
    border: 1px solid rgba(34,48,76,0.9) !important;
}
div[data-testid="stDataFrame"] th {
    background: rgba(15,23,42,0.9) !important;
    color: rgba(148,163,184,0.8) !important;
    font-size: 11px !important; letter-spacing: 1px !important;
}
div[data-testid="stDataFrame"] td {
    color: #e2e8f0 !important; font-size: 12px !important;
}

/* ── Filtros ──────────────────────────────────────────────── */
div[data-testid="stDateInput"] input,
div[data-testid="stTextInput"] input {
    background: rgba(20,23,40,0.9) !important;
    border: 1px solid rgba(255,255,255,0.1) !important;
    border-radius: 10px !important;
    color: #e2e8f0 !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 13px !important;
}
div[data-testid="stSelectbox"] > div > div,
div[data-testid="stMultiSelect"] > div > div {
    background: rgba(20,23,40,0.9) !important;
    border: 1px solid rgba(255,255,255,0.1) !important;
    border-radius: 10px !important;
    color: #e2e8f0 !important;
}
div[data-testid="stButton"] button {
    font-family: 'Bebas Neue', sans-serif !important;
    letter-spacing: 2px !important; font-size: 14px !important;
    border-radius: 10px !important; height: 44px !important;
    transition: all 0.15s !important;
}

/* ── Download button ──────────────────────────────────────── */
div[data-testid="stDownloadButton"] button {
    font-family: 'Bebas Neue', sans-serif !important;
    letter-spacing: 2px !important; font-size: 14px !important;
    border-radius: 10px !important; height: 48px !important;
    background: rgba(34,197,94,0.15) !important;
    border: 1px solid rgba(34,197,94,0.35) !important;
    color: #22c55e !important;
    width: 100% !important;
}
div[data-testid="stDownloadButton"] button:hover {
    background: rgba(34,197,94,0.25) !important;
}

/* ── Tabs ─────────────────────────────────────────────────── */
div[data-testid="stTabs"] [role="tablist"] {
    background: rgba(15,17,30,0.6) !important;
    border-radius: 12px !important; padding: 4px !important;
    border: 1px solid rgba(255,255,255,0.06) !important;
}
div[data-testid="stTabs"] [role="tab"] {
    font-family: 'Bebas Neue', sans-serif !important;
    letter-spacing: 2px !important; font-size: 14px !important;
    color: rgba(226,232,240,0.4) !important;
    border-radius: 8px !important; border: none !important;
}
div[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
    background: rgba(251,191,36,0.12) !important;
    color: #fbbf24 !important;
    border: 1px solid rgba(251,191,36,0.25) !important;
}

/* Scrollbar */
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: rgba(255,255,255,0.1); border-radius: 4px; }
</style>
"""

MESES_ES = {
    1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
    7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"
}

# ─── DB helpers ───────────────────────────────────────────────────────────────

def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def get_vendedores(distribuidor_id: int) -> List[str]:
    with get_conn() as c:
        rows = c.execute(
            """SELECT DISTINCT nombre_integrante FROM integrantes_grupo
               WHERE id_distribuidor = ? AND nombre_integrante IS NOT NULL
               ORDER BY nombre_integrante""",
            (distribuidor_id,)
        ).fetchall()
    return [r[0] for r in rows]


def get_tipos_pdv(distribuidor_id: int) -> List[str]:
    with get_conn() as c:
        rows = c.execute(
            """SELECT DISTINCT comentarios_telegram FROM exhibiciones
               WHERE id_distribuidor = ? AND comentarios_telegram IS NOT NULL
               ORDER BY comentarios_telegram""",
            (distribuidor_id,)
        ).fetchall()
    return [r[0] for r in rows if r[0]]


def query_exhibiciones(
    distribuidor_id: int,
    fecha_desde: date,
    fecha_hasta: date,
    vendedores: List[str],
    estados: List[str],
    tipos_pdv: List[str],
    nro_cliente: str,
) -> List[Dict]:
    wheres = [
        "e.id_distribuidor = ?",
        "DATE(e.timestamp_subida) >= ?",
        "DATE(e.timestamp_subida) <= ?",
    ]
    params: list = [distribuidor_id, fecha_desde.isoformat(), fecha_hasta.isoformat()]

    if vendedores:
        placeholders = ",".join("?" * len(vendedores))
        wheres.append(f"i.nombre_integrante IN ({placeholders})")
        params.extend(vendedores)

    if estados:
        placeholders = ",".join("?" * len(estados))
        wheres.append(f"e.estado IN ({placeholders})")
        params.extend(estados)

    if tipos_pdv:
        placeholders = ",".join("?" * len(tipos_pdv))
        wheres.append(f"e.comentarios_telegram IN ({placeholders})")
        params.extend(tipos_pdv)

    if nro_cliente.strip():
        wheres.append("e.numero_cliente_local LIKE ?")
        params.append(f"%{nro_cliente.strip()}%")

    sql = f"""
        SELECT
            e.id_exhibicion,
            i.nombre_integrante             AS vendedor,
            e.numero_cliente_local          AS cliente,
            e.comentarios_telegram          AS tipo_pdv,
            e.estado,
            e.supervisor_nombre             AS supervisor,
            e.comentarios                   AS comentario,
            e.timestamp_subida              AS fecha_carga,
            e.evaluated_at                  AS fecha_evaluacion,
            e.url_foto_drive                AS link_foto
        FROM exhibiciones e
        LEFT JOIN integrantes_grupo i ON i.id_integrante = e.id_integrante
        WHERE {" AND ".join(wheres)}
        ORDER BY e.timestamp_subida DESC
    """
    with get_conn() as c:
        rows = c.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


# ─── Excel export ─────────────────────────────────────────────────────────────

def build_excel(rows: List[Dict], filtros_desc: str) -> bytes:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exhibiciones"
    ws.sheet_view.showGridLines = False

    # Paleta
    BG_DARK   = "0B1220"
    BG_PANEL  = "121B2E"
    BG_HEADER = "0F172A"
    AMBER     = "FCD34D"
    CYAN      = "22D3EE"
    GREEN     = "22C55E"
    RED       = "EF4444"
    MUTED     = "94A3B8"
    WHITE     = "E2E8F0"

    estado_colors = {
        "Aprobado":  ("22C55E", "0F2E1A"),
        "Destacado": ("FCD34D", "2E2A0F"),
        "Rechazado": ("EF4444", "2E0F0F"),
        "Pendiente": ("94A3B8", "1A1F2E"),
    }

    thin = Side(style="thin", color="22304C")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(hex_color, bold=False, size=10):
        return Font(color=hex_color, bold=bold, size=size, name="Calibri")

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "CENTERMIND · REPORTE DE EXHIBICIONES"
    c.font  = Font(color=AMBER, bold=True, size=16, name="Calibri")
    c.fill  = fill(BG_DARK)
    c.alignment = align("left", "center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = filtros_desc
    c.font  = Font(color=MUTED, size=9, name="Calibri")
    c.fill  = fill(BG_DARK)
    c.alignment = align("left", "center")
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:J3")
    ws["A3"].fill = fill(BG_DARK)
    ws.row_dimensions[3].height = 8

    # ── Encabezados ───────────────────────────────────────────────────────────
    headers = ["ID", "VENDEDOR", "CLIENTE", "TIPO PDV", "ESTADO",
               "SUPERVISOR", "COMENTARIO", "FECHA CARGA", "FECHA EVALUACIÓN", "LINK FOTO"]
    col_widths = [8, 22, 14, 20, 14, 20, 30, 18, 18, 40]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font      = Font(color=AMBER, bold=True, size=9, name="Calibri")
        cell.fill      = fill(BG_HEADER)
        cell.alignment = align("center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[4].height = 22

    # ── Filas de datos ────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, 5):
        estado = row.get("estado") or "Pendiente"
        fg, bg = estado_colors.get(estado, (WHITE, BG_PANEL))

        def fmt_dt(val):
            if not val:
                return ""
            try:
                dt = datetime.fromisoformat(str(val))
                return dt.strftime("%d/%m/%Y %H:%M")
            except Exception:
                return str(val)[:16]

        values = [
            row.get("id_exhibicion"),
            row.get("vendedor") or "—",
            row.get("cliente") or "—",
            row.get("tipo_pdv") or "—",
            estado,
            row.get("supervisor") or "—",
            row.get("comentario") or "",
            fmt_dt(row.get("fecha_carga")),
            fmt_dt(row.get("fecha_evaluacion")),
            row.get("link_foto") or "",
        ]

        row_bg = "0D1526" if row_idx % 2 == 0 else BG_PANEL

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = align("left", "center", wrap=(col_idx == 7))

            if col_idx == 5:  # Estado con color propio
                cell.font = Font(color=fg, bold=True, size=9, name="Calibri")
                cell.fill = fill(bg)
            else:
                cell.font = Font(color=WHITE, size=9, name="Calibri")
                cell.fill = fill(row_bg)

        ws.row_dimensions[row_idx].height = 18

    # ── Fila total ────────────────────────────────────────────────────────────
    total_row = len(rows) + 5
    ws.merge_cells(f"A{total_row}:D{total_row}")
    c = ws[f"A{total_row}"]
    c.value = f"TOTAL: {len(rows)} registros"
    c.font  = Font(color=AMBER, bold=True, size=10, name="Calibri")
    c.fill  = fill(BG_DARK)
    c.alignment = align("right")

    # Congelar la fila de encabezados
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Gráficos ─────────────────────────────────────────────────────────────────

def render_charts(rows: List[Dict]):
    if not rows:
        return

    try:
        import pandas as pd
    except ImportError:
        st.warning("Instalá pandas para ver los gráficos: pip install pandas")
        return

    df = pd.DataFrame(rows)

    tab1, tab2, tab3 = st.tabs(["ESTADOS", "VENDEDORES", "TIMELINE"])

    # ── Tab 1: Distribución por estado ────────────────────────────────────────
    with tab1:
        st.markdown('<div class="card"><div class="card-title">Distribución por Estado</div>', unsafe_allow_html=True)
        estado_counts = df["estado"].value_counts().reset_index()
        estado_counts.columns = ["Estado", "Cantidad"]

        color_map = {
            "Aprobado":  "#22C55E",
            "Destacado": "#FCD34D",
            "Rechazado": "#EF4444",
            "Pendiente": "#94A3B8",
        }

        try:
            import plotly.express as px
            fig = px.pie(
                estado_counts, values="Cantidad", names="Estado",
                color="Estado",
                color_discrete_map=color_map,
                hole=0.45,
            )
            fig.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#e2e8f0", family="DM Sans"),
                legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#e2e8f0")),
                margin=dict(l=20, r=20, t=20, b=20),
            )
            fig.update_traces(textfont_color="#e2e8f0", textfont_size=13)
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            st.bar_chart(estado_counts.set_index("Estado"))

        st.markdown('</div>', unsafe_allow_html=True)

    # ── Tab 2: Por vendedor ────────────────────────────────────────────────────
    with tab2:
        st.markdown('<div class="card"><div class="card-title">Rendimiento por Vendedor</div>', unsafe_allow_html=True)

        vend_df = df.groupby(["vendedor", "estado"]).size().reset_index(name="cantidad")
        pivot   = vend_df.pivot_table(index="vendedor", columns="estado", values="cantidad", fill_value=0)

        for col in ["Aprobado", "Destacado", "Rechazado", "Pendiente"]:
            if col not in pivot.columns:
                pivot[col] = 0

        pivot = pivot[["Aprobado", "Destacado", "Rechazado", "Pendiente"]]
        pivot = pivot.sort_values("Aprobado", ascending=False).head(15)

        try:
            import plotly.graph_objects as go
            fig = go.Figure()
            bar_colors = {
                "Aprobado":  "#22C55E",
                "Destacado": "#FCD34D",
                "Rechazado": "#EF4444",
                "Pendiente": "#94A3B8",
            }
            for col in pivot.columns:
                fig.add_trace(go.Bar(
                    name=col,
                    x=pivot.index,
                    y=pivot[col],
                    marker_color=bar_colors[col],
                    marker_line_width=0,
                ))
            fig.update_layout(
                barmode="stack",
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#e2e8f0", family="DM Sans"),
                legend=dict(bgcolor="rgba(0,0,0,0)"),
                xaxis=dict(tickfont=dict(size=11), gridcolor="rgba(255,255,255,0.05)"),
                yaxis=dict(gridcolor="rgba(255,255,255,0.05)"),
                margin=dict(l=20, r=20, t=20, b=80),
            )
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            st.bar_chart(pivot)

        st.markdown('</div>', unsafe_allow_html=True)

    # ── Tab 3: Timeline ────────────────────────────────────────────────────────
    with tab3:
        st.markdown('<div class="card"><div class="card-title">Cargas por Día</div>', unsafe_allow_html=True)

        df["fecha"] = pd.to_datetime(df["fecha_carga"]).dt.date
        timeline = df.groupby(["fecha", "estado"]).size().reset_index(name="cantidad")

        try:
            import plotly.express as px
            color_map = {
                "Aprobado":  "#22C55E",
                "Destacado": "#FCD34D",
                "Rechazado": "#EF4444",
                "Pendiente": "#94A3B8",
            }
            fig = px.bar(
                timeline, x="fecha", y="cantidad", color="estado",
                color_discrete_map=color_map,
                barmode="stack",
            )
            fig.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#e2e8f0", family="DM Sans"),
                legend=dict(bgcolor="rgba(0,0,0,0)"),
                xaxis=dict(gridcolor="rgba(255,255,255,0.05)"),
                yaxis=dict(gridcolor="rgba(255,255,255,0.05)"),
                margin=dict(l=20, r=20, t=20, b=40),
            )
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            st.bar_chart(timeline.pivot_table(index="fecha", columns="estado", values="cantidad", fill_value=0))

        st.markdown('</div>', unsafe_allow_html=True)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    st.markdown(STYLE, unsafe_allow_html=True)

    u             = st.session_state.user
    dist_id       = u["id_distribuidor"]
    dist_nombre   = u.get("nombre_empresa", "").upper()
    usuario_login = u.get("usuario_login", "")

    # ── Topbar ────────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div class="topbar">
        <div style="display:flex;align-items:center;gap:20px;">
            <span class="topbar-logo">CENTERMIND · REPORTES</span>
            <span class="topbar-meta">{dist_nombre}</span>
        </div>
        <span class="user-badge">&#x1F4CA; {usuario_login}</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='padding:20px 24px 0;'>", unsafe_allow_html=True)

    # ── Panel de filtros ──────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-title">Filtros</div>', unsafe_allow_html=True)

    hoy      = date.today()
    mes_ini  = hoy.replace(day=1)

    # Fila 1: fechas + cliente
    c1, c2, c3, c4 = st.columns([2, 2, 2, 2])
    with c1:
        fecha_desde = st.date_input("Desde", value=mes_ini, key="f_desde",
                                     format="DD/MM/YYYY")
    with c2:
        fecha_hasta = st.date_input("Hasta", value=hoy, key="f_hasta",
                                     format="DD/MM/YYYY")
    with c3:
        nro_cliente = st.text_input("N° Cliente", placeholder="ej: 333", key="f_cliente")
    with c4:
        # Atajos de período
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        c4a, c4b, c4c = st.columns(3)
        with c4a:
            if st.button("HOY", key="p_hoy", use_container_width=True):
                st.session_state["f_desde"] = hoy
                st.session_state["f_hasta"] = hoy
                st.rerun()
        with c4b:
            if st.button("MES", key="p_mes", use_container_width=True):
                st.session_state["f_desde"] = mes_ini
                st.session_state["f_hasta"] = hoy
                st.rerun()
        with c4c:
            if st.button("TODO", key="p_todo", use_container_width=True):
                st.session_state["f_desde"] = date(2020, 1, 1)
                st.session_state["f_hasta"] = hoy
                st.rerun()

    # Fila 2: vendedor, estado, tipo PDV
    vendedores_opts = get_vendedores(dist_id)
    tipos_pdv_opts  = get_tipos_pdv(dist_id)
    estados_opts    = ["Pendiente", "Aprobado", "Destacado", "Rechazado"]

    c5, c6, c7 = st.columns([3, 3, 3])
    with c5:
        sel_vendedores = st.multiselect("Vendedor", vendedores_opts, key="f_vendedores",
                                         placeholder="Todos")
    with c6:
        sel_estados    = st.multiselect("Estado", estados_opts, key="f_estados",
                                         placeholder="Todos")
    with c7:
        sel_tipos      = st.multiselect("Tipo PDV", tipos_pdv_opts, key="f_tipos",
                                         placeholder="Todos")

    # Botón buscar
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    _, col_btn, _ = st.columns([4, 2, 4])
    with col_btn:
        buscar = st.button("BUSCAR", key="btn_buscar", use_container_width=True)

    st.markdown('</div>', unsafe_allow_html=True)

    # ── Ejecutar consulta ─────────────────────────────────────────────────────
    if "reporte_rows" not in st.session_state:
        st.session_state["reporte_rows"] = None

    if buscar:
        rows = query_exhibiciones(
            dist_id, fecha_desde, fecha_hasta,
            sel_vendedores, sel_estados, sel_tipos, nro_cliente
        )
        st.session_state["reporte_rows"] = rows
        st.session_state["reporte_filtros"] = {
            "desde": fecha_desde, "hasta": fecha_hasta,
            "vendedores": sel_vendedores, "estados": sel_estados,
            "tipos": sel_tipos, "cliente": nro_cliente,
        }

    rows = st.session_state.get("reporte_rows")

    if rows is None:
        st.markdown("""
        <div style="text-align:center;padding:60px 0;color:rgba(226,232,240,0.2);">
            <div style="font-family:'Bebas Neue',sans-serif;font-size:28px;letter-spacing:3px;">
                CONFIGURÁ LOS FILTROS Y PRESIONÁ BUSCAR
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)
        if st.button("← VOLVER AL MENU", key="btn_volver"):
            st.switch_page("app.py")
        return

    # ── KPIs del resultado ────────────────────────────────────────────────────
    total      = len(rows)
    aprobadas  = sum(1 for r in rows if r["estado"] == "Aprobado")
    destacadas = sum(1 for r in rows if r["estado"] == "Destacado")
    rechazadas = sum(1 for r in rows if r["estado"] == "Rechazado")
    pendientes = sum(1 for r in rows if r["estado"] == "Pendiente")
    puntos     = aprobadas + destacadas * 2

    filtros_info = st.session_state.get("reporte_filtros", {})
    desde_str = filtros_info.get("desde", fecha_desde)
    hasta_str = filtros_info.get("hasta", fecha_hasta)

    st.markdown(f"""
    <div class="card">
        <div class="card-title">Resultados &middot; {desde_str} al {hasta_str}</div>
        <div class="kpi-row">
            <div class="kpi-chip kpi-cyan">
                <div class="kpi-chip-num">{total}</div>
                <div class="kpi-chip-lbl">Total</div>
            </div>
            <div class="kpi-chip kpi-green">
                <div class="kpi-chip-num">{aprobadas}</div>
                <div class="kpi-chip-lbl">Aprobadas</div>
            </div>
            <div class="kpi-chip kpi-amber">
                <div class="kpi-chip-num">{destacadas}</div>
                <div class="kpi-chip-lbl">Destacadas</div>
            </div>
            <div class="kpi-chip kpi-red">
                <div class="kpi-chip-num">{rechazadas}</div>
                <div class="kpi-chip-lbl">Rechazadas</div>
            </div>
            <div class="kpi-chip kpi-muted">
                <div class="kpi-chip-num">{pendientes}</div>
                <div class="kpi-chip-lbl">Pendientes</div>
            </div>
            <div class="kpi-chip kpi-amber">
                <div class="kpi-chip-num">{puntos}</div>
                <div class="kpi-chip-lbl">Puntos</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if total == 0:
        st.markdown("""
        <div style="text-align:center;padding:40px 0;color:rgba(226,232,240,0.3);">
            <div style="font-family:'Bebas Neue',sans-serif;font-size:22px;letter-spacing:2px;">
                SIN RESULTADOS PARA LOS FILTROS SELECCIONADOS
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)
        if st.button("← VOLVER AL MENU", key="btn_volver2"):
            st.switch_page("app.py")
        return

    # ── Tabla + Exportar ──────────────────────────────────────────────────────
    col_tabla, col_export = st.columns([5, 1])

    with col_tabla:
        st.markdown('<div class="card"><div class="card-title">Tabla de Datos</div>', unsafe_allow_html=True)

        if HAS_PANDAS:
            import pandas as pd
            df = pd.DataFrame(rows)

            # Formatear fechas
            for col in ["fecha_carga", "fecha_evaluacion"]:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d/%m/%Y %H:%M")

            # Renombrar columnas para display
            col_rename = {
                "id_exhibicion":   "ID",
                "vendedor":        "Vendedor",
                "cliente":         "Cliente",
                "tipo_pdv":        "Tipo PDV",
                "estado":          "Estado",
                "supervisor":      "Supervisor",
                "comentario":      "Comentario",
                "fecha_carga":     "Fecha Carga",
                "fecha_evaluacion":"Fecha Eval.",
                "link_foto":       "Link Foto",
            }
            df = df.rename(columns=col_rename)
            # Ocultar link_foto de la tabla visual
            cols_show = [c for c in df.columns if c != "Link Foto"]
            st.dataframe(df[cols_show], use_container_width=True, height=420)
        else:
            # Fallback sin pandas
            for r in rows[:200]:
                st.markdown(
                    f'<div style="font-size:12px;color:#e2e8f0;padding:4px 0;border-bottom:1px solid rgba(255,255,255,0.04);">'
                    f'{r.get("vendedor","—")} | {r.get("cliente","—")} | {r.get("tipo_pdv","—")} | {r.get("estado","—")}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        st.markdown('</div>', unsafe_allow_html=True)

    with col_export:
        st.markdown('<div class="card"><div class="card-title">Exportar</div>', unsafe_allow_html=True)

        if HAS_OPENPYXL:
            filtros_sel = st.session_state.get("reporte_filtros", {})
            desde_e = filtros_sel.get("desde", fecha_desde)
            hasta_e = filtros_sel.get("hasta", fecha_hasta)
            filtros_desc = (
                f"Distribuidora: {dist_nombre} | "
                f"Período: {desde_e} al {hasta_e} | "
                f"Total: {total} registros | "
                f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            )
            excel_bytes = build_excel(rows, filtros_desc)
            filename    = f"centermind_{dist_nombre.lower().replace(' ','_')}_{desde_e}_{hasta_e}.xlsx"

            st.download_button(
                label="⬇ DESCARGAR EXCEL",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel",
                use_container_width=True,
            )
            st.markdown(
                f'<div style="font-size:10px;color:rgba(148,163,184,0.5);margin-top:8px;text-align:center;">'
                f'{total} filas · .xlsx</div>',
                unsafe_allow_html=True,
            )
        else:
            st.warning("Instalá openpyxl:\npip install openpyxl")

        st.markdown('</div>', unsafe_allow_html=True)

    # ── Gráficos ──────────────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-title">Gráficos</div>', unsafe_allow_html=True)
    render_charts(rows)
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Navegación ────────────────────────────────────────────────────────────
    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    if st.button("← VOLVER AL MENU", key="btn_volver3"):
        st.switch_page("app.py")

    st.markdown("</div>", unsafe_allow_html=True)


main()
