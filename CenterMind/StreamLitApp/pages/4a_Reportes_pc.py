# -*- coding: utf-8 -*-
"""
ShelfMind — Reportes
=====================
Filtros: fecha, vendedor, estado, tipo PDV, número de cliente.
Vista previa en tabla + gráficos + exportar a Excel con branding.
"""

from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional

import streamlit as st

# ─── Guard de sesión ──────────────────────────────────────────────────────────
if "logged_in" not in st.session_state or not st.session_state.logged_in:
    st.switch_page("app.py")

# ─── Configuración de página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="ShelfMind · Reportes",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── Imports opcionales ───────────────────────────────────────────────────────
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── CSS ──────────────────────────────────────────────────────────────────────
STYLE = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');

/* ── Paleta ShelfMind (Tobacco/Amber) ─────────────────────── */
:root {
    --bg-darkest:   #140E0C;
    --bg-card:      rgba(42, 30, 24, 0.5);
    --bg-input:     rgba(20, 15, 12, 0.8);
    
    --accent-amber: #D9A76A;
    --accent-hover: #FBBF24;
    --accent-sand:  #D9BD9C;
    
    --text-primary: #F0E6D8;
    --text-muted:   rgba(240, 230, 216, 0.5);
    --border-soft:  rgba(217, 167, 106, 0.15);
    
    /* Colores de estado ShelfMind */
    --st-aprobado:  #7DAF6B;
    --st-destacado: #FBBF24;
    --st-rechazado: #C0584A;
    --st-pendiente: #A68B72;
}

*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

html, body, [data-testid="stAppViewContainer"], [data-testid="stMain"] {
    background: var(--bg-darkest) !important;
    color: var(--text-primary) !important;
    font-family: 'DM Sans', sans-serif !important;
}

[data-testid="stHeader"], [data-testid="stToolbar"], [data-testid="stDecoration"] { display: none !important; }
[data-testid="stMainBlockContainer"]{ padding: 0 !important; max-width: 100% !important; }
section[data-testid="stSidebar"] { display: none !important; }

/* Fondo textura sutil */
[data-testid="stAppViewContainer"]::before {
    content: ''; position: fixed; inset: 0; z-index: 0;
    background:
        radial-gradient(ellipse 80% 50% at 10% 20%, rgba(217,167,106,0.04) 0%, transparent 60%),
        repeating-linear-gradient(0deg, transparent, transparent 39px, rgba(255,255,255,0.008) 40px),
        repeating-linear-gradient(90deg, transparent, transparent 39px, rgba(255,255,255,0.008) 40px);
    pointer-events: none;
}

/* ── Topbar ───────────────────────────────────────────────── */
.topbar {
    display: flex; align-items: center; justify-content: space-between;
    padding: 14px 32px;
    background: rgba(20, 14, 12, 0.95);
    border-bottom: 1px solid var(--border-soft);
    backdrop-filter: blur(8px);
    position: sticky; top: 0; z-index: 100;
}
.topbar-logo {
    font-family: 'Bebas Neue', sans-serif;
    font-size: 26px; letter-spacing: 3px; color: var(--accent-amber);
    text-shadow: 0 0 20px rgba(217, 167, 106, 0.3);
}
.topbar-meta { font-size: 12px; color: var(--text-muted); letter-spacing: 1px; }
.user-badge {
    display: inline-flex; align-items: center; gap: 6px;
    padding: 4px 14px; border-radius: 999px;
    font-size: 11px; letter-spacing: 1px;
    background: rgba(217, 167, 106, 0.1);
    border: 1px solid rgba(217, 167, 106, 0.25);
    color: var(--accent-amber);
}

/* ── Cards (Efecto Vidrio) ────────────────────────────────── */
.card {
    background: var(--bg-card);
    border: 1px solid var(--border-soft);
    border-radius: 16px;
    backdrop-filter: blur(12px);
    padding: 20px 22px;
    margin-bottom: 16px;
}
.card-title {
    font-size: 10px; letter-spacing: 3px; text-transform: uppercase;
    color: var(--accent-amber); margin-bottom: 16px;
    display: flex; align-items: center; gap: 8px; font-weight: 600;
}
.card-title::after { content: ''; flex: 1; height: 1px; background: var(--border-soft); }

/* ── KPI Chips ────────────────────────────────────────────── */
.kpi-row {
    display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 4px;
}
.kpi-chip {
    flex: 1; min-width: 110px;
    display: flex; flex-direction: column; align-items: center; justify-content: center;
    padding: 16px 10px; border-radius: 12px;
    background: rgba(20, 14, 12, 0.6);
    border: 1px solid var(--border-soft);
    transition: transform 0.2s ease;
}
.kpi-chip:hover { transform: translateY(-2px); }
.kpi-chip-num { font-family: 'Bebas Neue', sans-serif; font-size: 36px; line-height: 1; margin-bottom: 4px; }
.kpi-chip-lbl { font-size: 9px; letter-spacing: 2px; text-transform: uppercase; color: var(--text-muted); font-weight: 600; }

.kpi-total  .kpi-chip-num { color: var(--text-primary); }
.kpi-green  .kpi-chip-num { color: var(--st-aprobado); text-shadow: 0 0 10px rgba(125,175,107,0.3); }
.kpi-amber  .kpi-chip-num { color: var(--st-destacado); text-shadow: 0 0 10px rgba(251,191,36,0.3); }
.kpi-red    .kpi-chip-num { color: var(--st-rechazado); }
.kpi-muted  .kpi-chip-num { color: var(--st-pendiente); }

/* ── Filtros y Form Elements ──────────────────────────────── */
div[data-testid="stDateInput"] input,
div[data-testid="stTextInput"] input,
div[data-testid="stSelectbox"] > div > div,
div[data-testid="stMultiSelect"] > div > div {
    background: var(--bg-input) !important;
    border: 1px solid rgba(217, 167, 106, 0.2) !important;
    border-radius: 10px !important;
    color: var(--text-primary) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 13px !important;
    min-height: 44px !important;
}
div[data-testid="stDateInput"] input:focus,
div[data-testid="stTextInput"] input:focus,
div[data-testid="stSelectbox"] > div > div:focus-within,
div[data-testid="stMultiSelect"] > div > div:focus-within {
    border-color: var(--accent-amber) !important;
    box-shadow: 0 0 0 3px rgba(217, 167, 106, 0.1) !important;
}

/* ── Botones ──────────────────────────────────────────────── */
div[data-testid="stButton"] button {
    font-family: 'Bebas Neue', sans-serif !important;
    letter-spacing: 2px !important; font-size: 15px !important;
    border-radius: 10px !important; height: 44px !important;
    background: rgba(217, 167, 106, 0.1) !important;
    border: 1px solid rgba(217, 167, 106, 0.3) !important;
    color: var(--accent-amber) !important;
    transition: all 0.2s ease !important;
}
div[data-testid="stButton"] button:hover {
    background: rgba(217, 167, 106, 0.2) !important;
    border-color: var(--accent-hover) !important;
    color: var(--accent-hover) !important;
    transform: translateY(-2px) !important;
}

/* Botón Exportar GIGANTE */
div[data-testid="stDownloadButton"] button {
    font-family: 'Bebas Neue', sans-serif !important;
    letter-spacing: 2px !important; font-size: 18px !important;
    border-radius: 12px !important; height: 56px !important;
    background: rgba(125, 175, 107, 0.15) !important;
    border: 1px solid rgba(125, 175, 107, 0.4) !important;
    color: var(--st-aprobado) !important;
    width: 100% !important;
    box-shadow: 0 4px 15px rgba(125, 175, 107, 0.1) !important;
}
div[data-testid="stDownloadButton"] button:hover {
    background: rgba(125, 175, 107, 0.25) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 20px rgba(125, 175, 107, 0.2) !important;
}

/* ── Dataframes (Tabla) ───────────────────────────────────── */
div[data-testid="stDataFrame"] {
    border-radius: 12px !important;
    border: 1px solid var(--border-soft) !important;
}
div[data-testid="stDataFrame"] th {
    background: rgba(20, 14, 12, 0.9) !important;
    color: var(--accent-sand) !important;
    font-size: 11px !important; letter-spacing: 1px !important;
}
div[data-testid="stDataFrame"] td {
    color: var(--text-primary) !important; font-size: 12px !important;
}

/* ── Tabs (Gráficos) ──────────────────────────────────────── */
div[data-testid="stTabs"] [role="tablist"] {
    background: rgba(20, 14, 12, 0.6) !important;
    border-radius: 12px !important; padding: 4px !important;
    border: 1px solid var(--border-soft) !important;
}
div[data-testid="stTabs"] [role="tab"] {
    font-family: 'Bebas Neue', sans-serif !important;
    letter-spacing: 2px !important; font-size: 15px !important;
    color: var(--text-dim) !important;
    border-radius: 8px !important; border: none !important;
}
div[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
    background: rgba(217, 167, 106, 0.15) !important;
    color: var(--accent-amber) !important;
    border: 1px solid var(--border-soft) !important;
}

/* ── Scrollbars nativos ───────────────────────────────────── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: rgba(217, 167, 106, 0.2); border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: rgba(217, 167, 106, 0.4); }

/* ── Mobile Stacking ──────────────────────────────────────── */
@media (max-width: 768px) {
    div[data-testid="column"] { 
        width: 100% !important; 
        min-width: 100% !important; 
        padding: 0 !important; 
        margin-bottom: 8px !important;
    }
    .topbar { padding: 12px 16px; }
    .kpi-chip { min-width: 30%; } /* Caben 3 por fila en móvil */
}
</style>
"""

MESES_ES = {
    1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
    7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"
}

# ─── DB helpers ───────────────────────────────────────────────────────────────

# ─── API helpers ──────────────────────────────────────────────────────────────

def _api_conf():
    try:
        return st.secrets["API_URL"].rstrip("/"), st.secrets["API_KEY"]
    except Exception:
        return "http://localhost:8000", "shelfmind-clave-2025"

def _api_get(path: str):
    try:
        import requests
        url, key = _api_conf()
        r = requests.get(f"{url}{path}", headers={"x-api-key": key}, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception:
        return None

def _api_post(path: str, body: dict):
    try:
        import requests
        url, key = _api_conf()
        r = requests.post(f"{url}{path}", json=body, headers={"x-api-key": key}, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception:
        return None

def get_vendedores(distribuidor_id: int) -> List[str]:
    data = _api_get(f"/reportes/vendedores/{distribuidor_id}")
    return data if isinstance(data, list) else []

def get_tipos_pdv(distribuidor_id: int) -> List[str]:
    data = _api_get(f"/reportes/tipos-pdv/{distribuidor_id}")
    return data if isinstance(data, list) else []

def query_exhibiciones(
    distribuidor_id: int,
    fecha_desde: date,
    fecha_hasta: date,
    vendedores: List[str],
    estados: List[str],
    tipos_pdv: List[str],
    nro_cliente: str,
) -> List[Dict]:
    body = {
        "fecha_desde":  fecha_desde.isoformat(),
        "fecha_hasta":  fecha_hasta.isoformat(),
        "vendedores":   vendedores,
        "estados":      estados,
        "tipos_pdv":    tipos_pdv,
        "nro_cliente":  nro_cliente,
    }
    data = _api_post(f"/reportes/exhibiciones/{distribuidor_id}", body)
    return data if isinstance(data, list) else []

# ─── Excel export (Branding Premium) ──────────────────────────────────────────

def build_excel(rows: List[Dict], filtros_desc: str) -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte ShelfMind"
    ws.sheet_view.showGridLines = False

    # Paleta ShelfMind para Excel
    BG_DARK   = "140E0C"
    BG_PANEL  = "1A1311"
    BG_HEADER = "2A1E18"
    AMBER     = "D9A76A"
    GREEN     = "7DAF6B"
    RED       = "C0584A"
    MUTED     = "B8A392"
    WHITE     = "F0E6D8"
    
    estado_colors = {
        "Aprobado":  (GREEN, "1D2B1A"),
        "Destacado": (AMBER, "2E2310"),
        "Rechazado": (RED,   "2E1614"),
        "Pendiente": (MUTED, "1A1715"),
    }

    thin = Side(style="thin", color="4A3326")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def font(hex_color, bold=False, size=10): return Font(color=hex_color, bold=bold, size=size, name="Segoe UI")
    def align(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "SHELFMIND · REPORTE DE EXHIBICIONES"
    c.font  = font(AMBER, bold=True, size=16)
    c.fill  = fill(BG_DARK)
    c.alignment = align("left", "center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = filtros_desc
    c.font  = font(MUTED, size=9)
    c.fill  = fill(BG_DARK)
    c.alignment = align("left", "center")
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:J3")
    ws["A3"].fill = fill(BG_DARK)
    ws.row_dimensions[3].height = 8

    # ── Encabezados ───────────────────────────────────────────────────────────
    headers = ["ID", "VENDEDOR", "CLIENTE", "TIPO PDV", "ESTADO",
               "SUPERVISOR", "COMENTARIO", "FECHA CARGA", "FECHA EVALUACIÓN", "LINK FOTO"]
    col_widths = [8, 22, 14, 20, 14, 20, 30, 18, 18, 40]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font      = font(AMBER, bold=True, size=9)
        cell.fill      = fill(BG_HEADER)
        cell.alignment = align("center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[4].height = 22

    # ── Filas de datos ────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, 5):
        estado = row.get("estado") or "Pendiente"
        fg, bg = estado_colors.get(estado, (WHITE, BG_PANEL))

        def fmt_dt(val):
            if not val: return ""
            try: return datetime.fromisoformat(str(val)).strftime("%d/%m/%Y %H:%M")
            except Exception: return str(val)[:16]

        values = [
            row.get("id_exhibicion"),
            row.get("vendedor") or "—",
            row.get("cliente") or "—",
            row.get("tipo_pdv") or "—",
            estado,
            row.get("supervisor") or "—",
            row.get("comentario") or "",
            fmt_dt(row.get("fecha_carga")),
            fmt_dt(row.get("fecha_evaluacion")),
            row.get("link_foto") or "",
        ]

        row_bg = "18110D" if row_idx % 2 == 0 else BG_PANEL

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = align("left", "center", wrap=(col_idx == 7))

            if col_idx == 5:  # Estado con color propio
                cell.font = font(fg, bold=True, size=9)
                cell.fill = fill(bg)
            else:
                cell.font = font(WHITE, size=9)
                cell.fill = fill(row_bg)

        ws.row_dimensions[row_idx].height = 18

    # ── Fila total ────────────────────────────────────────────────────────────
    total_row = len(rows) + 5
    ws.merge_cells(f"A{total_row}:D{total_row}")
    c = ws[f"A{total_row}"]
    c.value = f"TOTAL: {len(rows)} registros"
    c.font  = font(AMBER, bold=True, size=10)
    c.fill  = fill(BG_DARK)
    c.alignment = align("right")

    ws.freeze_panes = "A5"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Gráficos ─────────────────────────────────────────────────────────────────

def render_charts(rows: List[Dict]):
    if not rows: return

    try:
        import pandas as pd
    except ImportError:
        st.warning("Instalá pandas para ver los gráficos: pip install pandas")
        return

    df = pd.DataFrame(rows)
    tab1, tab2, tab3 = st.tabs(["ESTADOS", "VENDEDORES", "TIMELINE"])

    # Paleta ShelfMind para Plotly
    color_map = {
        "Aprobado":  "#7DAF6B",
        "Destacado": "#FBBF24",
        "Rechazado": "#C0584A",
        "Pendiente": "#A68B72",
    }

    # ── Tab 1: Distribución por estado ────────────────────────────────────────
    with tab1:
        st.markdown('<div class="card"><div class="card-title">Distribución por Estado</div>', unsafe_allow_html=True)
        estado_counts = df["estado"].value_counts().reset_index()
        estado_counts.columns = ["Estado", "Cantidad"]

        try:
            import plotly.express as px
            fig = px.pie(
                estado_counts, values="Cantidad", names="Estado",
                color="Estado", color_discrete_map=color_map, hole=0.45,
            )
            fig.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#F0E6D8", family="DM Sans"),
                legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#F0E6D8")),
                margin=dict(l=20, r=20, t=20, b=20),
            )
            fig.update_traces(textfont_color="#1A1311", textfont_size=14)
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            st.bar_chart(estado_counts.set_index("Estado"))
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Tab 2: Por vendedor ────────────────────────────────────────────────────
    with tab2:
        st.markdown('<div class="card"><div class="card-title">Rendimiento por Vendedor</div>', unsafe_allow_html=True)
        vend_df = df.groupby(["vendedor", "estado"]).size().reset_index(name="cantidad")
        pivot   = vend_df.pivot_table(index="vendedor", columns="estado", values="cantidad", fill_value=0)

        for col in ["Aprobado", "Destacado", "Rechazado", "Pendiente"]:
            if col not in pivot.columns: pivot[col] = 0

        pivot = pivot[["Aprobado", "Destacado", "Rechazado", "Pendiente"]]
        pivot = pivot.sort_values("Aprobado", ascending=False).head(15)

        try:
            import plotly.graph_objects as go
            fig = go.Figure()
            for col in pivot.columns:
                fig.add_trace(go.Bar(
                    name=col, x=pivot.index, y=pivot[col],
                    marker_color=color_map[col], marker_line_width=0,
                ))
            fig.update_layout(
                barmode="stack",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#F0E6D8", family="DM Sans"),
                legend=dict(bgcolor="rgba(0,0,0,0)"),
                xaxis=dict(tickfont=dict(size=11), gridcolor="rgba(217, 167, 106, 0.1)"),
                yaxis=dict(gridcolor="rgba(217, 167, 106, 0.1)"),
                margin=dict(l=20, r=20, t=20, b=80),
            )
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            st.bar_chart(pivot)
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Tab 3: Timeline ────────────────────────────────────────────────────────
    with tab3:
        st.markdown('<div class="card"><div class="card-title">Cargas por Día</div>', unsafe_allow_html=True)
        df["fecha"] = pd.to_datetime(df["fecha_carga"]).dt.date
        timeline = df.groupby(["fecha", "estado"]).size().reset_index(name="cantidad")

        try:
            import plotly.express as px
            fig = px.bar(
                timeline, x="fecha", y="cantidad", color="estado",
                color_discrete_map=color_map, barmode="stack",
            )
            fig.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#F0E6D8", family="DM Sans"),
                legend=dict(bgcolor="rgba(0,0,0,0)"),
                xaxis=dict(gridcolor="rgba(217, 167, 106, 0.1)"),
                yaxis=dict(gridcolor="rgba(217, 167, 106, 0.1)"),
                margin=dict(l=20, r=20, t=20, b=40),
            )
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            st.bar_chart(timeline.pivot_table(index="fecha", columns="estado", values="cantidad", fill_value=0))
        st.markdown('</div>', unsafe_allow_html=True)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    st.markdown(STYLE, unsafe_allow_html=True)

    u             = st.session_state.user
    dist_id       = u["id_distribuidor"]
    dist_nombre   = u.get("nombre_empresa", "").upper()
    usuario_login = u.get("usuario_login", "")

    # ── Topbar ────────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div class="topbar">
        <div style="display:flex;align-items:center;gap:20px;">
            <span class="topbar-logo">SHELFMIND · REPORTES</span>
            <span class="topbar-meta">{dist_nombre}</span>
        </div>
        <span class="user-badge">&#x1F4CA; {usuario_login}</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='padding:20px 24px 0;'>", unsafe_allow_html=True)

    # ── Panel de filtros ──────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-title">Filtros Analíticos</div>', unsafe_allow_html=True)

    hoy      = date.today()
    mes_ini  = hoy.replace(day=1)

    # Fila 1: fechas + cliente
    c1, c2, c3, c4 = st.columns([2, 2, 2, 2])
    with c1: fecha_desde = st.date_input("Desde", value=mes_ini, key="f_desde", format="DD/MM/YYYY")
    with c2: fecha_hasta = st.date_input("Hasta", value=hoy, key="f_hasta", format="DD/MM/YYYY")
    with c3: nro_cliente = st.text_input("N° Cliente", placeholder="ej: 333", key="f_cliente")
    with c4:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        c4a, c4b, c4c = st.columns(3)
        with c4a:
            if st.button("HOY", key="p_hoy", use_container_width=True):
                st.session_state["f_desde"] = hoy
                st.session_state["f_hasta"] = hoy
                st.rerun()
        with c4b:
            if st.button("MES", key="p_mes", use_container_width=True):
                st.session_state["f_desde"] = mes_ini
                st.session_state["f_hasta"] = hoy
                st.rerun()
        with c4c:
            if st.button("TODO", key="p_todo", use_container_width=True):
                st.session_state["f_desde"] = date(2020, 1, 1)
                st.session_state["f_hasta"] = hoy
                st.rerun()

    # Fila 2: vendedor, estado, tipo PDV
    vendedores_opts = get_vendedores(dist_id)
    tipos_pdv_opts  = get_tipos_pdv(dist_id)
    estados_opts    = ["Pendiente", "Aprobado", "Destacado", "Rechazado"]

    c5, c6, c7 = st.columns([3, 3, 3])
    with c5: sel_vendedores = st.multiselect("Vendedor", vendedores_opts, key="f_vendedores", placeholder="Todos")
    with c6: sel_estados    = st.multiselect("Estado", estados_opts, key="f_estados", placeholder="Todos")
    with c7: sel_tipos      = st.multiselect("Tipo PDV", tipos_pdv_opts, key="f_tipos", placeholder="Todos")

    # Botón buscar
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    _, col_btn, _ = st.columns([4, 2, 4])
    with col_btn:
        buscar = st.button("BUSCAR", key="btn_buscar", use_container_width=True)

    st.markdown('</div>', unsafe_allow_html=True)

    # ── Ejecutar consulta ─────────────────────────────────────────────────────
    if "reporte_rows" not in st.session_state:
        st.session_state["reporte_rows"] = None

    if buscar:
        rows = query_exhibiciones(
            dist_id, fecha_desde, fecha_hasta,
            sel_vendedores, sel_estados, sel_tipos, nro_cliente
        )
        st.session_state["reporte_rows"] = rows
        st.session_state["reporte_filtros"] = {
            "desde": fecha_desde, "hasta": fecha_hasta,
            "vendedores": sel_vendedores, "estados": sel_estados,
            "tipos": sel_tipos, "cliente": nro_cliente,
        }

    rows = st.session_state.get("reporte_rows")

    if rows is None:
        st.markdown("""
        <div style="text-align:center;padding:60px 0;color:rgba(217, 167, 106, 0.3);">
            <div style="font-family:'Bebas Neue',sans-serif;font-size:28px;letter-spacing:3px;">
                CONFIGURÁ LOS FILTROS Y PRESIONÁ BUSCAR
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)
        if st.button("← VOLVER AL MENU", key="btn_volver"): st.switch_page("app.py")
        return

    # ── KPIs del resultado ────────────────────────────────────────────────────
    total      = len(rows)
    aprobadas  = sum(1 for r in rows if r["estado"] == "Aprobado")
    destacadas = sum(1 for r in rows if r["estado"] == "Destacado")
    rechazadas = sum(1 for r in rows if r["estado"] == "Rechazado")
    pendientes = sum(1 for r in rows if r["estado"] == "Pendiente")
    puntos     = aprobadas + destacadas * 2

    filtros_info = st.session_state.get("reporte_filtros", {})
    desde_str = filtros_info.get("desde", fecha_desde)
    hasta_str = filtros_info.get("hasta", fecha_hasta)

    st.markdown(f"""
    <div class="card">
        <div class="card-title">Resultados &middot; {desde_str} al {hasta_str}</div>
        <div class="kpi-row">
            <div class="kpi-chip kpi-total">
                <div class="kpi-chip-num">{total}</div>
                <div class="kpi-chip-lbl">Total</div>
            </div>
            <div class="kpi-chip kpi-green">
                <div class="kpi-chip-num">{aprobadas}</div>
                <div class="kpi-chip-lbl">Aprobadas</div>
            </div>
            <div class="kpi-chip kpi-amber">
                <div class="kpi-chip-num">{destacadas}</div>
                <div class="kpi-chip-lbl">Destacadas</div>
            </div>
            <div class="kpi-chip kpi-red">
                <div class="kpi-chip-num">{rechazadas}</div>
                <div class="kpi-chip-lbl">Rechazadas</div>
            </div>
            <div class="kpi-chip kpi-muted">
                <div class="kpi-chip-num">{pendientes}</div>
                <div class="kpi-chip-lbl">Pendientes</div>
            </div>
            <div class="kpi-chip kpi-amber">
                <div class="kpi-chip-num">{puntos}</div>
                <div class="kpi-chip-lbl">Puntos</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if total == 0:
        st.markdown("""
        <div style="text-align:center;padding:40px 0;color:rgba(217, 167, 106, 0.4);">
            <div style="font-family:'Bebas Neue',sans-serif;font-size:22px;letter-spacing:2px;">
                SIN RESULTADOS PARA LOS FILTROS SELECCIONADOS
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)
        if st.button("← VOLVER AL MENU", key="btn_volver2"): st.switch_page("app.py")
        return

    # ── Tabla + Exportar (Stacked en móvil, Lado a lado en PC) ────────────────
    
    # Renderizamos botón Exportar primero (Por UX móvil: descargar > ver tabla chiquita)
    if HAS_OPENPYXL:
        filtros_sel = st.session_state.get("reporte_filtros", {})
        desde_e = filtros_sel.get("desde", fecha_desde)
        hasta_e = filtros_sel.get("hasta", fecha_hasta)
        filtros_desc = (
            f"Distribuidora: {dist_nombre} | "
            f"Período: {desde_e} al {hasta_e} | "
            f"Total: {total} registros | "
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        )
        excel_bytes = build_excel(rows, filtros_desc)
        filename    = f"shelfmind_{dist_nombre.lower().replace(' ','_')}_{desde_e}_{hasta_e}.xlsx"

        st.markdown('<div class="card" style="padding: 16px 22px;">', unsafe_allow_html=True)
        c_vacio, c_btn, c_vacio2 = st.columns([1, 2, 1])
        with c_btn:
            st.download_button(
                label="⬇ DESCARGAR REPORTE EXCEL",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel",
                use_container_width=True,
            )
            st.markdown(
                f'<div style="font-size:11px;color:var(--text-muted);margin-top:8px;text-align:center;letter-spacing:1px;">'
                f'Genera archivo Excel optimizado · {total} filas</div>',
                unsafe_allow_html=True,
            )
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.warning("Instalá openpyxl para descargar en Excel: pip install openpyxl")

    # Renderizamos Tabla
    st.markdown('<div class="card"><div class="card-title">Vista Previa de Datos</div>', unsafe_allow_html=True)
    if HAS_PANDAS:
        df = pd.DataFrame(rows)
        for col in ["fecha_carga", "fecha_evaluacion"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d/%m/%Y %H:%M")
        col_rename = {
            "id_exhibicion":   "ID",
            "vendedor":        "Vendedor",
            "cliente":         "Cliente",
            "tipo_pdv":        "Tipo PDV",
            "estado":          "Estado",
            "supervisor":      "Supervisor",
            "comentario":      "Comentario",
            "fecha_carga":     "Fecha Carga",
            "fecha_evaluacion":"Fecha Eval.",
            "link_foto":       "Link Foto",
        }
        df = df.rename(columns=col_rename)
        cols_show = [c for c in df.columns if c != "Link Foto"]
        st.dataframe(df[cols_show], use_container_width=True, height=400)
    else:
        for r in rows[:200]:
            st.markdown(
                f'<div style="font-size:12px;color:var(--text-primary);padding:6px 0;border-bottom:1px solid var(--border-soft);">'
                f'{r.get("vendedor","—")} | {r.get("cliente","—")} | {r.get("tipo_pdv","—")} | {r.get("estado","—")}'
                f'</div>',
                unsafe_allow_html=True,
            )
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Gráficos ──────────────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-title">Gráficos Analíticos</div>', unsafe_allow_html=True)
    render_charts(rows)
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Navegación ────────────────────────────────────────────────────────────
    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    if st.button("← VOLVER AL MENU", key="btn_volver3"):
        st.switch_page("app.py")

    st.markdown("</div>", unsafe_allow_html=True)


if __name__ == "__main__":
    main()