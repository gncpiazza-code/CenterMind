# file: cuentascorrientes.py
from __future__ import annotations

import os
import re
import threading
import datetime as dt
from typing import Dict, List, Optional
import pandas as pd
import numpy as np
import unicodedata
from tkinter import messagebox

# Importar matplotlib para generar el PNG del gráfico
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker # Para formatear el dinero en el gráfico

# utils.py es necesario para leer_excel. Asegúrate de que esté en la misma carpeta.
from utils import leer_excel

__all__ = ["procesar_cuentas_corrientes"]

# --- INICIO DE LA MODIFICACIÓN: Mapa de Sucursales ---
SUCURSALES_MAP = {
    "1": "Reconquista",
    "2": "Resistencia",
    "3": "Saenz Peña",
    "4": "Corrientes",
    "5": "Cordoba",
}
# --- FIN DE LA MODIFICACIÓN ---

# ----------------------------- Utilidades -----------------------------
def _strip_accents(text: str) -> str:
    if not isinstance(text, str): return ""
    return "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))

def _norm(s: str) -> str:
    s = _strip_accents(str(s)).lower().strip()
    s = re.sub(r"[\W_]+", " ", s)
    return re.sub(r"\s+", " ", s)

def _first_match(colnames: List[str], patterns: List[str]) -> str | None:
    norm_map = {c: _norm(c) for c in colnames}
    norm_patterns = [_norm(p) for p in patterns]
    for c, nc in norm_map.items():
        if any(p in nc for p in norm_patterns):
            return c
    return None

def _plot_antiguedad_png(analysis_df: pd.DataFrame, out_png_path: str, vendedor_nombre: str) -> str:
    if analysis_df.empty or analysis_df['saldo_total'].sum() == 0:
        return ""

    labels = analysis_df.index
    sizes = analysis_df['saldo_total'].fillna(0) # Asegurarse de que no haya NaN
    total_saldo = sizes.sum()
    
    # === CAMBIO DE COLORES ===
    colors = ['#34a853', "#fa9c0f", "#F32b26", '#4285f4', '#000000']
    # ==========================
    
    # Ajustar el tamaño de la figura para dar espacio a la leyenda
    fig, ax = plt.subplots(figsize=(11, 7)) # Más ancho que alto
    
    # 1. Dibujar el gráfico de dona SIN etiquetas internas (autopct=None, labels=None)
    pie_wedges = ax.pie(sizes,
                        colors=colors,
                        startangle=90,
                        wedgeprops=dict(width=0.4, edgecolor='w'),
                        labels=None,
                        autopct=None) # Quitamos el autopct

    # 2. Mantener el título y el aspecto
    ax.set_title(f'Distribución de Deuda por Antigüedad - {vendedor_nombre}', fontsize=12, fontweight='bold', pad=20)
    ax.axis('equal')

    # 3. Generar las etiquetas para la leyenda
    legend_labels = []
    for i in range(len(labels)):
        label_text = str(labels[i])
        value = sizes.iloc[i]
        
        # Calcular porcentaje
        if total_saldo > 0:
            percent = (value / total_saldo)
            percent_text = f"{percent:.1%}"
        else:
            percent_text = "0.0%"
            
        # Formatear dinero
        money_text = f"${value:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
        
        # Crear la etiqueta completa
        full_label = f"{label_text}: {percent_text} ({money_text})"
        legend_labels.append(full_label)

    # 4. Añadir la leyenda fuera del gráfico
    ax.legend(pie_wedges[0],
              legend_labels,
              title="Rangos de Antigüedad",
              loc="upper left", # <--- CAMBIO DE POSICIÓN
              bbox_to_anchor=(0.95, 0.9), # <--- CAMBIO DE POSICIÓN (más arriba)
              fontsize=12, # <--- CAMBIO DE TAMAÑO
              frameon=False, # Sin borde para un look más limpio
              labelspacing=1.7) # <--- CAMBIO DE INTERLINEADO

    # No usar plt.tight_layout(), ya que bbox_inches='tight' en savefig lo maneja mejor
    try:
        # Guardar la figura, 'bbox_inches_tight' es clave para incluir la leyenda externa
        plt.savefig(out_png_path, dpi=150, bbox_inches='tight')
    except Exception as e:
        plt.close(fig)
        raise IOError(f"No se pudo guardar el gráfico en '{out_png_path}'. Verifique los permisos de la carpeta. Error: {e}")
    
    plt.close(fig)
    return out_png_path

def insertar_png_en_hoja(hoja, png_path: str, celda: str = "A1", width: Optional[int] = None, height: Optional[int] = None):
    if not png_path or not os.path.exists(png_path):
        print(f"Advertencia: No se encontró el archivo de imagen '{png_path}' para insertar en el Excel.")
        return
    from openpyxl.drawing.image import Image as XLImage
    img = XLImage(png_path)
    
    if width:
        img.width = width
    if height:
        img.height = height
        
    hoja.add_image(img, celda)

# ----------------------------- Lógica Principal del Reporte -----------------------------
CANONICAL = {
    "sucursal": ["sucursal"],
    "vendedor": ["vendedor"],
    "cliente": ["cliente"],
    "antiguedad": ["antiguedad deuda", "antiguedad", "antigüedad"],
    "cant_cbte": ["cant cbte", "cantidad comprobantes"],
    "saldo_total": ["saldo total"],
}

def map_columns(df: pd.DataFrame) -> Dict[str, str | None]:
    cols = list(df.columns)
    mapping: Dict[str, str | None] = {}
    for key, patterns in CANONICAL.items():
        found = _first_match(cols, patterns)
        if not found:
            raise ValueError(f"No se encontró una columna para '{key}'. Se buscó: {patterns}")
        mapping[key] = found
    return mapping

def generar_reporte_cuentas_corrientes(path_excel: str, out_dir: str) -> str:
    subfolder_path = os.path.join(out_dir, "Cuentas Corrientes")
    os.makedirs(subfolder_path, exist_ok=True)
    
    df_raw = leer_excel(path_excel)
    
    mapping = map_columns(df_raw)

    mapping_inverso = {v: k for k, v in mapping.items()}
    df = df_raw.rename(columns=mapping_inverso)

    # --- INICIO DE LA MODIFICACIÓN: Traducir código de sucursal ---
    # Nos aseguramos que la columna sea texto, quitamos espacios
    df['sucursal'] = df['sucursal'].astype(str).str.strip()
    # Traducimos usando el mapa. Si no encuentra (ej: "99" o "Chaco"), 
    # .fillna(df['sucursal']) mantiene el valor original.
    df['sucursal'] = df['sucursal'].map(SUCURSALES_MAP).fillna(df['sucursal'])
    # --- FIN DE LA MODIFICACIÓN ---

    df['vendedor'] = df['vendedor'].fillna("SIN VENDEDOR").astype(str)
    df = df[~df['vendedor'].str.contains("SIN VENDEDOR", case=False, na=False)]

    for col in ["cant_cbte", "saldo_total", "antiguedad"]:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    sucursales = df["sucursal"].dropna().unique()
    sucursal_nombre = str(sucursales[0]) if len(sucursales) > 0 else "Sucursal"

    fecha_hoy_str = dt.datetime.now().strftime("%d-%m-%Y")
    
    out_name = f"Cuentas Corrientes {sucursal_nombre} {fecha_hoy_str}.xlsx"
    out_path = os.path.join(subfolder_path, out_name) # Guardar en la subcarpeta

    temp_excel_path = os.path.join(subfolder_path, f"temp_cc_report_{dt.datetime.now().timestamp()}.xlsx")
    
    png_insertion_map = {}

    with pd.ExcelWriter(temp_excel_path, engine="xlsxwriter") as writer:
        vendedores = sorted(df["vendedor"].unique())
        
        for vend in vendedores:
            df_vend = df[df["vendedor"] == vend].copy()
            
            df_vend_sorted = df_vend[df_vend["saldo_total"] > 0].sort_values(
                by=["antiguedad", "saldo_total"], ascending=[False, False]
            )

            if df_vend_sorted.empty:
                continue
            
            output_df = df_vend_sorted[["vendedor", "cliente", "cant_cbte", "saldo_total", "antiguedad"]].rename(columns={
                "vendedor": "Vendedor",
                "cliente": "Cliente",
                "cant_cbte": "Cant. Comprobantes",
                "saldo_total": "Saldo Total",
                "antiguedad": "Antigüedad (días)"
            })
            
            sheet_name = re.sub(r'[\\/*?:\[\]]', '', str(vend))[:31]
            output_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            workbook = writer.book
            ws = writer.sheets[sheet_name]
            money_format = workbook.add_format({'num_format': '$#,##0.00'})
            header_format = workbook.add_format({'bold': True, 'text_wrap': True, 'valign': 'top', 'fg_color': '#D7E4BC', 'border': 1, 'align': 'center'})
            bold_format = workbook.add_format({'bold': True})
            percent_format = workbook.add_format({'num_format': '0.0%'})
            
            for col_num, value in enumerate(output_df.columns.values):
                ws.write(0, col_num, value, header_format)

            ws.set_column('A:B', 35); ws.set_column('C:C', 18); ws.set_column('D:D', 15, money_format); ws.set_column('E:E', 18)
            ws.freeze_panes(1, 0)

            num_rows_data = len(output_df)
            total_row_idx = num_rows_data + 1
            ws.write(total_row_idx, 2, "TOTAL", bold_format)
            total_formula = f'=SUM(D2:D{total_row_idx})'
            ws.write_formula(total_row_idx, 3, total_formula, workbook.add_format({'bold': True, 'num_format': '$#,##0.00'}))

            # === CAMBIO DE INTERVALOS (BINS) Y ETIQUETAS ===
            bins = [-1, 7, 15, 21, 30, float('inf')]
            labels = ['1-7 Días', '8-15 Días', '16-21 Días', '22-30 Días', '+30 Días']
            # =================================================
            
            df_vend_sorted['rango_antiguedad'] = pd.cut(df_vend_sorted['antiguedad'], bins=bins, labels=labels, right=True)

            analysis = df_vend_sorted.groupby('rango_antiguedad', observed=False).agg(
                saldo_total=('saldo_total', 'sum'),
                cant_clientes=('cliente', 'count')
            ).reindex(labels)
            
            total_clientes = analysis['cant_clientes'].sum()
            analysis['porc_clientes'] = (analysis['cant_clientes'] / total_clientes) if total_clientes > 0 else 0
            
            analysis_start_row = 1
            ws.merge_range(analysis_start_row, 6, analysis_start_row, 8, "Análisis por Antigüedad", header_format)
            
            headers = ["Rango", "% Clientes", "Saldo Total"]
            for col, header in enumerate(headers):
                ws.write(analysis_start_row + 1, 6 + col, header, bold_format)

            current_row = analysis_start_row + 2
            for rango, row_data in analysis.iterrows():
                ws.write(current_row, 6, rango); ws.write(current_row, 7, row_data['porc_clientes'], percent_format); ws.write(current_row, 8, row_data['saldo_total'], money_format)
                current_row += 1
            
            ws.set_column('G:G', 15); ws.set_column('H:H', 12); ws.set_column('I:I', 15)

            if not analysis.empty and analysis['saldo_total'].sum() > 0:
                safe_name = re.sub(r'[^\w\s-]', '', str(vend)).strip()[:40]
                png_out = os.path.join(subfolder_path, f"deuda_antiguedad_{safe_name}.png") # Guardar PNG en subcarpeta
                png_out = os.path.normpath(png_out)
                png_path = _plot_antiguedad_png(analysis, png_out, str(vend))
                
                if png_path:
                    chart_cell = f'A{total_row_idx + 3}'
                    png_insertion_map[sheet_name] = {'path': png_path, 'cell': chart_cell}
    
    from openpyxl import load_workbook
    wb = load_workbook(temp_excel_path)
    
    for sheet_name, info in png_insertion_map.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            png_path = info['path']
            png_cell = info['cell']
            insertar_png_en_hoja(ws, png_path, celda=png_cell, width=600, height=420)
            
    wb.save(out_path)
    
    os.remove(temp_excel_path)
    for info in png_insertion_map.values():
        png_path_to_delete = info['path']
        if os.path.exists(png_path_to_delete):
            os.remove(png_path_to_delete)
    
    return out_path

# =================== ADAPTADOR PARA LA INTERFAZ GRÁFICA =====================
# NOTA: Este bloque fue reemplazado por la lógica Web (FastAPI) en `services/cuentas_corrientes_service.py`
# Ya no se utiliza Tkinter ni MessageBox aquí.
