import os
import re
import datetime as dt
import pandas as pd
import numpy as np
import unicodedata
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import load_workbook
import json

# Local utility imports
from utils import leer_excel

# Re-use the SUCURSALES_MAP and canonical mappings from legacy
SUCURSALES_MAP = {
    "1": "Reconquista",
    "2": "Resistencia",
    "3": "Saenz Peña",
    "4": "Corrientes",
    "5": "Cordoba",
}

CANONICAL = {
    "sucursal": ["sucursal"],
    "vendedor": ["vendedor"],
    "cliente": ["cliente"],
    "antiguedad": ["antiguedad deuda", "antiguedad", "antigüedad"],
    "cant_cbte": ["cant cbte", "cantidad comprobantes"],
    "saldo_total": ["saldo total"],
}

# --- Utils Strings --- 
def _strip_accents(text: str) -> str:
    if not isinstance(text, str): return ""
    return "".join(ch for ch in unicodedata.normalize("NFKD", str(text)) if not unicodedata.combining(ch))

def _norm(s: str) -> str:
    s = _strip_accents(str(s)).lower().strip()
    s = re.sub(r"[\W_]+", " ", s)
    return re.sub(r"\s+", " ", s)

def _first_match(colnames: list[str], patterns: list[str]) -> str | None:
    norm_map = {c: _norm(c) for c in colnames}
    norm_patterns = [_norm(p) for p in patterns]
    for c, nc in norm_map.items():
        if any(p in nc for p in norm_patterns):
            return c
    return None

def map_columns(df: pd.DataFrame) -> dict[str, str | None]:
    cols = list(df.columns)
    mapping: dict[str, str | None] = {}
    for key, patterns in CANONICAL.items():
        found = _first_match(cols, patterns)
        if not found:
            raise ValueError(f"No se encontró una columna para '{key}'. Se buscó: {patterns}")
        mapping[key] = found
    return mapping

# --- Matplotlib Chart Generation ---
def _plot_antiguedad_png(analysis_df: pd.DataFrame, out_png_path: str, vendedor_nombre: str) -> str:
    if analysis_df.empty or analysis_df['saldo_total'].sum() == 0:
        return ""

    labels = analysis_df.index
    sizes = analysis_df['saldo_total'].fillna(0)
    total_saldo = sizes.sum()
    
    colors = ['#34a853', "#fa9c0f", "#F32b26", '#4285f4', '#000000']
    
    fig, ax = plt.subplots(figsize=(11, 7))
    
    pie_wedges = ax.pie(sizes,
                        colors=colors,
                        startangle=90,
                        wedgeprops=dict(width=0.4, edgecolor='w'),
                        labels=None,
                        autopct=None)

    ax.set_title(f'Distribución de Deuda por Antigüedad - {vendedor_nombre}', fontsize=12, fontweight='bold', pad=20)
    ax.axis('equal')

    legend_labels = []
    for i in range(len(labels)):
        label_text = str(labels[i])
        value = sizes.iloc[i]
        
        if total_saldo > 0:
            percent = (value / total_saldo)
            percent_text = f"{percent:.1%}"
        else:
            percent_text = "0.0%"
            
        money_text = f"${value:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
        full_label = f"{label_text}: {percent_text} ({money_text})"
        legend_labels.append(full_label)

    ax.legend(pie_wedges[0],
              legend_labels,
              title="Rangos de Antigüedad",
              loc="upper left",
              bbox_to_anchor=(0.95, 0.9),
              fontsize=12,
              frameon=False,
              labelspacing=1.7)

    try:
        plt.savefig(out_png_path, dpi=150, bbox_inches='tight')
    except Exception as e:
        plt.close(fig)
        raise IOError(f"No se pudo guardar el gráfico en '{out_png_path}'. Error: {e}")
    
    plt.close(fig)
    return out_png_path

def insertar_png_en_hoja(hoja, png_path: str, celda: str = "A1", width: int | None = None, height: int | None = None):
    if not png_path or not os.path.exists(png_path):
        return
    from openpyxl.drawing.image import Image as XLImage
    img = XLImage(png_path)
    
    if width:
        img.width = width
    if height:
        img.height = height
        
    hoja.add_image(img, celda)

# --- Logica de Alertas (MOTOR) ---
def aplicar_alertas_credito(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """Añade la columna 'Alerta de Crédito' según las reglas y excepciones."""
    reglas = config.get("reglas_generales", {})
    excepciones = {
        _norm(exc.get("cliente", "")): exc
        for exc in config.get("excepciones", [])
    }
    
    def evaluar_fila(row):
        cliente_norm = _norm(row["cliente"])
        
        # 1. Determinar qué bloque de reglas usamos
        r_dinero = reglas.get("limite_dinero", {"activo": False})
        r_cbte = reglas.get("limite_cbte", {"activo": False})
        r_dias = reglas.get("limite_dias", {"activo": False})
        
        if cliente_norm in excepciones:
            exc = excepciones[cliente_norm]
            if "limite_dinero" in exc: r_dinero = exc["limite_dinero"]
            if "limite_cbte" in exc: r_cbte = exc["limite_cbte"]
            if "limite_dias" in exc: r_dias = exc["limite_dias"]

        motivos = []
        
        # 2. Evaluar condiciones
        if r_dinero.get("activo") and row["saldo_total"] > r_dinero.get("valor", float('inf')):
            motivos.append(f"Dinero (>{r_dinero['valor']})")
            
        if r_cbte.get("activo") and row["cant_cbte"] > r_cbte.get("valor", float('inf')):
            motivos.append(f"Cbtes (>{r_cbte['valor']})")
            
        if r_dias.get("activo") and row["antiguedad"] > r_dias.get("valor", float('inf')):
            motivos.append(f"Días (>{r_dias['valor']})")
            
        return "⚠️ Excedido en: " + " | ".join(motivos) if motivos else ""

    df["Alerta de Crédito"] = df.apply(evaluar_fila, axis=1)
    return df

# --- Generador Principal ---
def procesar_cuentas_corrientes_service(path_excel: str, out_dir: str, config: dict):
    """
    Lee, procesa, aplica alertas y devuelve:
    (path_al_excel_resultado, json_data_para_previsualizar)
    """
    subfolder_path = os.path.join(out_dir, "Procesados")
    os.makedirs(subfolder_path, exist_ok=True)
    
    # 1. Leer archivo
    df_raw = leer_excel(path_excel)
    mapping = map_columns(df_raw)
    mapping_inverso = {v: k for k, v in mapping.items()}
    df = df_raw.rename(columns=mapping_inverso).copy()

    # 2. Limpieza de datos básica
    df['sucursal'] = df['sucursal'].astype(str).str.strip()
    df['sucursal'] = df['sucursal'].map(SUCURSALES_MAP).fillna(df['sucursal'])
    
    df['vendedor'] = df['vendedor'].fillna("SIN VENDEDOR").astype(str)
    df = df[~df['vendedor'].str.contains("SIN VENDEDOR", case=False, na=False)]

    for col in ["cant_cbte", "saldo_total", "antiguedad"]:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
    # 3. Inyectar Motor de Alertas
    df = aplicar_alertas_credito(df, config)

    # Variables
    sucursales = df["sucursal"].dropna().unique()
    sucursal_nombre = str(sucursales[0]) if len(sucursales) > 0 else "Sucursal"
    fecha_hoy_str = dt.datetime.now().strftime("%d-%m-%Y")
    
    out_name = f"Reporte_Ventas_{sucursal_nombre}_{fecha_hoy_str}_{dt.datetime.now().timestamp()}.xlsx"
    out_path = os.path.join(subfolder_path, out_name)
    temp_excel_path = os.path.join(subfolder_path, f"temp_{out_name}")
    
    png_insertion_map = {}
    json_data = {"resumen_alertas": [], "vendedores": {}}

    with pd.ExcelWriter(temp_excel_path, engine="xlsxwriter") as writer:
        workbook = writer.book
        money_format = workbook.add_format({'num_format': '$#,##0.00'})
        header_format = workbook.add_format({'bold': True, 'text_wrap': True, 'valign': 'top', 'fg_color': '#D7E4BC', 'border': 1, 'align': 'center'})
        bold_format = workbook.add_format({'bold': True})
        percent_format = workbook.add_format({'num_format': '0.0%'})
        
        # --- A) PESTAÑA NUEVA: Resumen Alertas ---
        df_alertas = df[df["Alerta de Crédito"] != ""].copy()
        df_alertas_sorted = df_alertas.sort_values(by="saldo_total", ascending=False)
        
        resumen_cols = ["vendedor", "cliente", "saldo_total", "cant_cbte", "antiguedad", "Alerta de Crédito"]
        output_resumen = df_alertas_sorted[resumen_cols].rename(columns={
            "vendedor": "Vendedor",
            "cliente": "Cliente",
            "saldo_total": "Saldo Total",
            "cant_cbte": "Cant. Cbtes",
            "antiguedad": "Antigüedad",
        })
        
        output_resumen.to_excel(writer, sheet_name="Resumen Alertas", index=False)
        ws_resumen = writer.sheets["Resumen Alertas"]
        
        for col_num, value in enumerate(output_resumen.columns.values):
            ws_resumen.write(0, col_num, value, header_format)
            
        ws_resumen.set_column('A:A', 20)
        ws_resumen.set_column('B:B', 35)
        ws_resumen.set_column('C:C', 15, money_format)
        ws_resumen.set_column('D:E', 12)
        ws_resumen.set_column('F:F', 35) # Alerta
        
        # Populate JSON for Resumen
        json_data["resumen_alertas"] = output_resumen.to_dict(orient="records")
        
        # --- B) PESTAÑAS DE VENDEDORES (Legacy adaptado) ---
        vendedores = sorted(df["vendedor"].unique())
        
        for vend in vendedores:
            df_vend = df[df["vendedor"] == vend].copy()
            df_vend_sorted = df_vend[df_vend["saldo_total"] > 0].sort_values(
                by=["antiguedad", "saldo_total"], ascending=[False, False]
            )

            if df_vend_sorted.empty:
                continue
            
            output_df = df_vend_sorted[["vendedor", "cliente", "cant_cbte", "saldo_total", "antiguedad", "Alerta de Crédito"]].rename(columns={
                "vendedor": "Vendedor",
                "cliente": "Cliente",
                "cant_cbte": "Cant. Comprobantes",
                "saldo_total": "Saldo Total",
                "antiguedad": "Antigüedad (días)"
            })
            
            sheet_name = re.sub(r'[\\/*?:\[\]]', '', str(vend))[:31]
            output_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            
            for col_num, value in enumerate(output_df.columns.values):
                ws.write(0, col_num, value, header_format)

            ws.set_column('A:B', 30); ws.set_column('C:C', 18); ws.set_column('D:D', 15, money_format); ws.set_column('E:E', 18); ws.set_column('F:F', 25)
            ws.freeze_panes(1, 0)

            num_rows_data = len(output_df)
            total_row_idx = num_rows_data + 1
            ws.write(total_row_idx, 2, "TOTAL", bold_format)
            total_formula = f'=SUM(D2:D{total_row_idx})'
            ws.write_formula(total_row_idx, 3, total_formula, workbook.add_format({'bold': True, 'num_format': '$#,##0.00'}))

            # Intervals Chart
            bins = [-1, 7, 15, 21, 30, float('inf')]
            labels = ['1-7 Días', '8-15 Días', '16-21 Días', '22-30 Días', '+30 Días']
            df_vend_sorted['rango_antiguedad'] = pd.cut(df_vend_sorted['antiguedad'], bins=bins, labels=labels, right=True)

            analysis = df_vend_sorted.groupby('rango_antiguedad', observed=False).agg(
                saldo_total=('saldo_total', 'sum'),
                cant_clientes=('cliente', 'count')
            ).reindex(labels)
            
            total_clientes = analysis['cant_clientes'].sum()
            analysis['porc_clientes'] = (analysis['cant_clientes'] / total_clientes) if total_clientes > 0 else 0
            
            analysis_start_row = 1
            ws.merge_range(analysis_start_row, 7, analysis_start_row, 9, "Análisis por Antigüedad", header_format)
            
            headers = ["Rango", "% Clientes", "Saldo Total"]
            for col, header in enumerate(headers):
                ws.write(analysis_start_row + 1, 7 + col, header, bold_format)

            current_row = analysis_start_row + 2
            for rango, row_data in analysis.iterrows():
                ws.write(current_row, 7, rango); ws.write(current_row, 8, row_data['porc_clientes'], percent_format); ws.write(current_row, 9, row_data['saldo_total'], money_format)
                current_row += 1
            
            ws.set_column('H:H', 15); ws.set_column('I:I', 12); ws.set_column('J:J', 15)

            if not analysis.empty and analysis['saldo_total'].sum() > 0:
                safe_name = re.sub(r'[^\w\s-]', '', str(vend)).strip()[:40]
                png_out = os.path.join(subfolder_path, f"deuda_antiguedad_{safe_name}.png")
                png_path = _plot_antiguedad_png(analysis, png_out, str(vend))
                
                if png_path:
                    chart_cell = f'A{total_row_idx + 3}'
                    png_insertion_map[sheet_name] = {'path': png_path, 'cell': chart_cell}

            # Populate JSON for Vendor
            # Resetting index to easily export to_dict
            json_data["vendedores"][vend] = {
                "tabla": output_df.to_dict(orient="records"),
                "grafico_analisis": analysis.reset_index().to_dict(orient="records")
            }
    
    # --- C) Insertar PNGs post-generación (con openpyxl) ---
    wb = load_workbook(temp_excel_path)
    for sheet_name, info in png_insertion_map.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            insertar_png_en_hoja(ws, info['path'], celda=info['cell'], width=600, height=420)
            
    wb.save(out_path)
    os.remove(temp_excel_path)
    
    # Limpiar PNGs
    for info in png_insertion_map.values():
        if os.path.exists(info['path']):
            os.remove(info['path'])
            
    return out_path, json_data
