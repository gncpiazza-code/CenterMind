# -*- coding: utf-8 -*-
"""
services/objetivos_liquidacion_service.py
==========================================
Liquidación monetaria de objetivos de compañía.

Tablas involucradas (ya migradas vía 20260605_objetivos_flags_liquidacion.sql):
  - objetivos_liquidacion_tarifas: tipo TEXT PK, monto_vendedor NUMERIC, activo BOOL, updated_at
  - objetivos_liquidacion_bono:    id INT PK (=1), bono_mando_medio NUMERIC, updated_at, updated_by UUID
  - objetivos.liquidacion_at:      TIMESTAMPTZ NULL (nueva columna en tabla objetivos)
"""
from __future__ import annotations

import io
import logging
from calendar import monthrange
from datetime import datetime, timezone, timedelta
from typing import Any

from db import sb
from core.tenant_tables import tenant_table_name

logger = logging.getLogger("LiquidacionService")

PAGE = 1000
_TARIFA_TABLE = "objetivos_liquidacion_tarifas"
_BONO_TABLE = "objetivos_liquidacion_bono"
_OBJETIVOS_TABLE = "objetivos"

# Tipo excluido de liquidación v1
_TIPOS_EXCLUIDOS = {"cobranza"}

_TIPO_LABELS: dict[str, str] = {
    "exhibicion": "Exhibición",
    "ruteo_alteo": "Alteo",
    "compradores": "Compradores",
    "conversion_estado": "Activación",
}

_FMT_ARS = '"$" #,##0.00'
_FMT_PCT = '0.0"%"'


# ─── Helpers internos ─────────────────────────────────────────────────────────

def _safe_float(val: Any, fallback: float = 0.0) -> float:
    try:
        return float(val)
    except (TypeError, ValueError):
        return fallback


def _safe_int(val: Any, fallback: int = 0) -> int:
    try:
        return int(val)
    except (TypeError, ValueError):
        return fallback


def _lazy_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        return openpyxl, Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl no instalado — no se puede generar XLSX")


def _header_row(ws, row: list[str], openpyxl_mods, row_num: int = 1):
    _, Font, PatternFill, Alignment = openpyxl_mods
    violet = "7C3AED"
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(fill_type="solid", fgColor=violet)
        cell.alignment = Alignment(horizontal="center")


def _subtitle_row(ws, text: str, openpyxl_mods, row_num: int, ncols: int = 11):
    _, Font, PatternFill, Alignment = openpyxl_mods
    gray = "D1D5DB"
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col_idx, value=text if col_idx == 1 else None)
        cell.fill = PatternFill(fill_type="solid", fgColor=gray)
        if col_idx == 1:
            cell.font = Font(bold=True, size=10)


def _autofit(ws, max_width: int = 50, start_row: int = 1):
    for col in ws.columns:
        cells = [c for c in col if c.row >= start_row]
        if not cells:
            continue
        max_len = max((len(str(c.value or "")) for c in cells), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, max_width)


def _mes_rango(mes_yyyy_mm: str) -> tuple[str, str]:
    """Retorna (primer_dia, ultimo_dia) del mes en formato YYYY-MM-DD."""
    year, month = int(mes_yyyy_mm[:4]), int(mes_yyyy_mm[5:7])
    last_day = monthrange(year, month)[1]
    return f"{mes_yyyy_mm}-01", f"{mes_yyyy_mm}-{last_day:02d}"


def _mes_label(mes_yyyy_mm: str) -> str:
    meses = (
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    )
    try:
        year, month = int(mes_yyyy_mm[:4]), int(mes_yyyy_mm[5:7])
        if 1 <= month <= 12:
            return f"{meses[month - 1]} {year}"
    except (ValueError, IndexError):
        pass
    return mes_yyyy_mm


def _tipo_label(tipo: str) -> str:
    return _TIPO_LABELS.get(tipo, tipo.replace("_", " ").title())


def _load_distribuidor_nombre(dist_id: int) -> str:
    try:
        for col in ("id_distribuidor", "id"):
            rows = (
                sb.table("distribuidores")
                .select("nombre_empresa")
                .eq(col, dist_id)
                .limit(1)
                .execute()
                .data
                or []
            )
            if rows:
                return str(rows[0].get("nombre_empresa") or f"Distribuidora {dist_id}")
    except Exception as exc:
        logger.warning("[liquidacion] nombre distribuidor dist=%s: %s", dist_id, exc)
    return f"Distribuidora {dist_id}"


def _fmt_fecha(val: Any) -> str:
    if not val:
        return ""
    s = str(val)[:10]
    if len(s) == 10 and s[4] == "-":
        return f"{s[8:10]}/{s[5:7]}/{s[0:4]}"
    return s


def _lazy_openpyxl_full():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl no instalado — no se puede generar XLSX")


def _xlsx_thin_border(Border, Side):
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _xlsx_write_header_row(ws, headers: list[str], row: int, mods, *, ncols: int | None = None):
    _, Font, PatternFill, Alignment, Border, Side, _ = mods
    border = _xlsx_thin_border(Border, Side)
    n = ncols or len(headers)
    for col, label in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(fill_type="solid", fgColor="5B21B6")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 22


def _xlsx_money(cell, value: float) -> None:
    cell.value = round(_safe_float(value), 2)
    cell.number_format = _FMT_ARS


def _xlsx_pct(cell, value: float) -> None:
    cell.value = round(_safe_float(value), 1)
    cell.number_format = _FMT_PCT


def _xlsx_autofit(ws, max_width: int = 44, min_width: int = 10) -> None:
    _, _, _, _, _, _, get_column_letter = _lazy_openpyxl_full()
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = max(min_width, min(max_len + 3, max_width))


def _paginate(query_fn) -> list[dict]:
    """Ejecuta query paginado PAGE=1000 y retorna todas las filas."""
    rows: list[dict] = []
    offset = 0
    while True:
        batch = query_fn(offset)
        rows.extend(batch or [])
        if len(batch or []) < PAGE:
            break
        offset += PAGE
    return rows


# ─── Configuración ────────────────────────────────────────────────────────────

def get_config() -> dict:
    """Lee tarifas y bono de mando medio de Supabase."""
    tarifas_rows = (
        sb.table(_TARIFA_TABLE)
        .select("tipo,monto_vendedor,activo,updated_at")
        .order("tipo")
        .execute()
        .data
        or []
    )

    bono_rows = (
        sb.table(_BONO_TABLE)
        .select("bono_mando_medio,updated_at")
        .eq("id", 1)
        .limit(1)
        .execute()
        .data
        or []
    )
    bono_row = bono_rows[0] if bono_rows else {}

    tarifas = [
        {
            "tipo": r["tipo"],
            "monto_vendedor": _safe_float(r.get("monto_vendedor")),
            "activo": bool(r.get("activo", True)),
        }
        for r in tarifas_rows
    ]

    return {
        "tarifas": tarifas,
        "bono_mando_medio": _safe_float(bono_row.get("bono_mando_medio")),
        "updated_at": str(bono_row.get("updated_at") or ""),
    }


def put_config(data, updated_by: str | None = None) -> dict:
    """
    Persiste tarifas y bono en Supabase.
    `data` debe ser LiquidacionConfigIn o un dict compatible.
    """
    now_iso = datetime.now(timezone.utc).isoformat()

    # Upsert tarifas
    for tarifa in (data.tarifas if hasattr(data, "tarifas") else data.get("tarifas", [])):
        row = {
            "tipo": tarifa.tipo if hasattr(tarifa, "tipo") else tarifa["tipo"],
            "monto_vendedor": (
                tarifa.monto_vendedor
                if hasattr(tarifa, "monto_vendedor")
                else tarifa["monto_vendedor"]
            ),
            "activo": tarifa.activo if hasattr(tarifa, "activo") else tarifa.get("activo", True),
            "updated_at": now_iso,
        }
        sb.table(_TARIFA_TABLE).upsert(row, on_conflict="tipo").execute()

    # Actualizar bono
    bono_mando = (
        data.bono_mando_medio
        if hasattr(data, "bono_mando_medio")
        else data.get("bono_mando_medio", 0.0)
    )
    bono_update: dict[str, Any] = {
        "id": 1,
        "bono_mando_medio": bono_mando,
        "updated_at": now_iso,
    }
    if updated_by:
        bono_update["updated_by"] = updated_by
    sb.table(_BONO_TABLE).upsert(bono_update, on_conflict="id").execute()

    return get_config()


# ─── Compute ──────────────────────────────────────────────────────────────────

def _load_tarifas_activas() -> dict[str, float]:
    """Retorna {tipo: monto_vendedor} solo para tarifas activas y no excluidas."""
    rows = (
        sb.table(_TARIFA_TABLE)
        .select("tipo,monto_vendedor")
        .eq("activo", True)
        .execute()
        .data
        or []
    )
    return {
        r["tipo"]: _safe_float(r["monto_vendedor"])
        for r in rows
        if r.get("tipo") not in _TIPOS_EXCLUIDOS
    }


def _load_vendedores_nombre(dist_id: int) -> dict[int, str]:
    """Carga {id_vendedor: nombre_erp} del tenant."""
    t_vend = tenant_table_name("vendedores_v2", dist_id)

    def _query(offset: int):
        return (
            sb.table(t_vend)
            .select("id_vendedor,nombre_erp")
            .range(offset, offset + PAGE - 1)
            .execute()
            .data
            or []
        )

    rows = _paginate(_query)
    return {
        _safe_int(r["id_vendedor"]): str(r.get("nombre_erp") or "")
        for r in rows
        if r.get("id_vendedor") is not None
    }


def _load_objetivos_compania_mes(dist_id: int, mes_inicio: str, mes_fin: str) -> list[dict]:
    """
    Carga todos los objetivos compañía del dist_id cuyo mes_referencia cae en el mes
    (o fallback fecha_objetivo). Incluye cumplidos y no-cumplidos para calcular mando medio.
    """
    # Cargamos todos los objetivos compañía del distribuidor en el mes
    # Filtramos por mes_referencia o fecha_objetivo dentro del mes
    def _query(offset: int):
        return (
            sb.table(_OBJETIVOS_TABLE)
            .select(
                "id,id_vendedor,tipo,descripcion,valor_objetivo,valor_actual,"
                "cumplido,resultado_final,mes_referencia,fecha_objetivo,"
                "created_at,desglose_cache"
            )
            .eq("id_distribuidor", dist_id)
            .eq("origen", "compania")
            .gte("mes_referencia", mes_inicio)
            .lte("mes_referencia", mes_fin)
            .range(offset, offset + PAGE - 1)
            .execute()
            .data
            or []
        )

    rows = _paginate(_query)

    # Filtrar tipos sin tarifa (incluyendo cobranza y tipos no activos)
    # En compute_liquidacion se filtra por tipo, aquí cargamos todo para mando medio
    return rows


def _load_objetivos_cumplidos_compania_mes(
    dist_id: int, mes_inicio: str, mes_fin: str, tipos_activos: set[str]
) -> list[dict]:
    """
    Carga objetivos compañía cumplidos (resultado_final='exito') del mes,
    solo para tipos con tarifa activa.
    """
    def _query(offset: int):
        return (
            sb.table(_OBJETIVOS_TABLE)
            .select(
                "id,id_vendedor,tipo,descripcion,valor_objetivo,valor_actual,"
                "cumplido,resultado_final,mes_referencia,fecha_objetivo,"
                "created_at,desglose_cache"
            )
            .eq("id_distribuidor", dist_id)
            .eq("origen", "compania")
            .eq("cumplido", True)
            .eq("resultado_final", "exito")
            .gte("mes_referencia", mes_inicio)
            .lte("mes_referencia", mes_fin)
            .in_("tipo", list(tipos_activos))
            .range(offset, offset + PAGE - 1)
            .execute()
            .data
            or []
        )

    return _paginate(_query)


def compute_liquidacion(dist_id: int, mes_yyyy_mm: str) -> dict:
    """
    Calcula la liquidación completa de objetivos de compañía para un mes.
    """
    mes_inicio, mes_fin = _mes_rango(mes_yyyy_mm)

    # 1. Cargar configuración
    config = get_config()
    tarifas_activas = _load_tarifas_activas()  # {tipo: monto}
    tipos_activos = set(tarifas_activas.keys())
    bono_mando_medio = _safe_float(config.get("bono_mando_medio"))

    # 2. Cargar nombres de vendedores
    nombres_vendedor = _load_vendedores_nombre(dist_id)

    # 3. Cargar objetivos cumplidos para detalle de vendedores
    objs_cumplidos = _load_objetivos_cumplidos_compania_mes(
        dist_id, mes_inicio, mes_fin, tipos_activos
    )

    # 4. Cargar TODOS los objetivos compañía del mes (para mando medio)
    objs_todos = _load_objetivos_compania_mes(dist_id, mes_inicio, mes_fin)
    # Filtrar por tipos activos (excluye cobranza y tipos sin tarifa)
    objs_todos_filtrados = [o for o in objs_todos if o.get("tipo") in tipos_activos]

    # 5. Construir detalle por vendedor×objetivo
    vendedor_rows: list[dict] = []
    for obj in objs_cumplidos:
        id_vend = _safe_int(obj.get("id_vendedor"))
        tipo = str(obj.get("tipo") or "")
        meta = _safe_float(obj.get("valor_objetivo"))
        avance = _safe_float(obj.get("valor_actual"))
        pct = min(100.0, round(avance / meta * 100, 1)) if meta > 0 else 100.0
        monto = tarifas_activas.get(tipo, 0.0)

        # avance_pdvs para exhibicion con min_pdvs_distintos
        desglose = obj.get("desglose_cache") or {}
        if isinstance(desglose, str):
            import json
            try:
                desglose = json.loads(desglose)
            except Exception:
                desglose = {}
        avance_pdvs: int | None = None
        if tipo == "exhibicion":
            pdvs_val = desglose.get("pdvs_distintos_count")
            if pdvs_val is not None:
                avance_pdvs = _safe_int(pdvs_val)

        vendedor_rows.append(
            {
                "id_vendedor": id_vend,
                "nombre_vendedor": nombres_vendedor.get(id_vend) or None,
                "id_objetivo": str(obj.get("id") or ""),
                "tipo": tipo,
                "tipo_label": _tipo_label(tipo),
                "descripcion": obj.get("descripcion") or None,
                "fecha_objetivo": _fmt_fecha(obj.get("fecha_objetivo")),
                "meta": meta,
                "avance": avance,
                "avance_pdvs": avance_pdvs,
                "pct": pct,
                "cumplido": True,
                "resultado_final": str(obj.get("resultado_final") or "exito"),
                "tarifa_unitaria": monto,
                "monto": monto,
            }
        )

    # Ordenar por nombre vendedor y luego tipo
    vendedor_rows.sort(
        key=lambda r: (str(r.get("nombre_vendedor") or "").lower(), r.get("tipo", ""))
    )

    # 6. Calcular mando medio
    # Agrupar todos los objetivos (filtrados por tipos activos) por vendedor
    from collections import defaultdict
    vend_objetivos: dict[int, list[dict]] = defaultdict(list)
    for obj in objs_todos_filtrados:
        id_vend = _safe_int(obj.get("id_vendedor"))
        vend_objetivos[id_vend].append(obj)

    mando_medio_rows: list[dict] = []
    for id_vend, objs_vend in vend_objetivos.items():
        asignados = len(objs_vend)
        cumplidos = sum(
            1
            for o in objs_vend
            if bool(o.get("cumplido")) and str(o.get("resultado_final") or "") == "exito"
        )
        # Factor: 0.5 si 1 asignado y 1 cumplido; 1.0 si todos cumplidos y >=2; 0 resto
        if asignados == 1 and cumplidos == 1:
            factor = 0.5
        elif cumplidos == asignados and asignados >= 2:
            factor = 1.0
        else:
            factor = 0.0
        monto_bono = round(factor * bono_mando_medio, 2)

        mando_medio_rows.append(
            {
                "id_vendedor": id_vend,
                "nombre_vendedor": nombres_vendedor.get(id_vend) or None,
                "asignados": asignados,
                "cumplidos": cumplidos,
                "factor": factor,
                "monto_bono": monto_bono,
            }
        )

    mando_medio_rows.sort(
        key=lambda r: str(r.get("nombre_vendedor") or "").lower()
    )

    # 7. Totales
    total_vendedores = round(sum(r["monto"] for r in vendedor_rows), 2)
    total_mando_medio = round(sum(r["monto_bono"] for r in mando_medio_rows), 2)
    total_distribuidora = round(total_vendedores + total_mando_medio, 2)

    return {
        "dist_id": dist_id,
        "mes": mes_yyyy_mm,
        "vendedores": vendedor_rows,
        "mando_medio": mando_medio_rows,
        "total_vendedores": total_vendedores,
        "total_mando_medio": total_mando_medio,
        "total_distribuidora": total_distribuidora,
    }


# ─── Export XLSX ──────────────────────────────────────────────────────────────

def export_xlsx(dist_id: int, mes_yyyy_mm: str) -> bytes:
    """
    Genera XLSX de liquidación — una sola hoja con bloques:
      cabecera + tarifas | detalle por objetivo | mando medio | totales.
    """
    from collections import OrderedDict

    liq = compute_liquidacion(dist_id, mes_yyyy_mm)
    config = get_config()
    dist_nombre = _load_distribuidor_nombre(dist_id)
    mes_inicio, mes_fin = _mes_rango(mes_yyyy_mm)
    mes_humano = _mes_label(mes_yyyy_mm)
    generado = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")

    mods = _lazy_openpyxl_full()
    openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter = mods
    border = _xlsx_thin_border(Border, Side)

    NCOLS = 13
    fill_section = PatternFill(fill_type="solid", fgColor="F3F4F6")
    fill_subtotal = PatternFill(fill_type="solid", fgColor="EDE9FE")
    fill_zebra = PatternFill(fill_type="solid", fgColor="FAFAFA")
    fill_total = PatternFill(fill_type="solid", fgColor="5B21B6")
    font_total = Font(bold=True, color="FFFFFF", size=11)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidación"
    ws.sheet_view.showGridLines = False

    r = 1

    def merge_title(text: str, *, size: int = 14, color: str = "5B21B6") -> None:
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=True, size=size, color=color)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

    def section_bar(text: str) -> None:
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=True, size=10, color="374151")
        c.fill = fill_section
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, NCOLS + 1):
            ws.cell(row=r, column=col).fill = fill_section
        ws.row_dimensions[r].height = 20
        r += 1

    def meta_pair(label: str, value: str) -> None:
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10, color="6B7280")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NCOLS)
        ws.cell(row=r, column=2, value=value).font = Font(size=10)
        r += 1

    # ── Cabecera ───────────────────────────────────────────────────────────────
    merge_title("LIQUIDACIÓN — OBJETIVOS COMPAÑÍA")
    r += 1
    meta_pair("Distribuidora", f"{dist_nombre} (ID {dist_id})")
    meta_pair("Mes de referencia", f"{mes_humano} ({mes_yyyy_mm})")
    meta_pair("Período", f"{_fmt_fecha(mes_inicio)} → {_fmt_fecha(mes_fin)}")
    meta_pair("Generado", generado)
    meta_pair(
        "Objetivos liquidados",
        f"{len(liq['vendedores'])} objetivo(s) · "
        f"{len({row['id_vendedor'] for row in liq['vendedores']})} vendedor(es) con pago",
    )
    r += 1

    # ── Tarifas vigentes ───────────────────────────────────────────────────────
    section_bar("TARIFAS GLOBALES VIGENTES")
    tarifa_headers = ["Tipo", "Tarifa vendedor", "Estado"]
    _xlsx_write_header_row(ws, tarifa_headers, r, mods)
    tarifa_header_row = r
    r += 1
    bono_base = _safe_float(config.get("bono_mando_medio"))
    for tarifa in config.get("tarifas") or []:
        if not tarifa.get("activo"):
            continue
        tipo = str(tarifa.get("tipo") or "")
        ws.cell(row=r, column=1, value=_tipo_label(tipo)).border = border
        c_tar = ws.cell(row=r, column=2)
        _xlsx_money(c_tar, tarifa.get("monto_vendedor"))
        c_tar.border = border
        ws.cell(row=r, column=3, value="Activo").border = border
        r += 1
    ws.cell(row=r, column=1, value="Bono mando medio (base)").font = Font(bold=True)
    ws.cell(row=r, column=1).border = border
    c_bono = ws.cell(row=r, column=2)
    _xlsx_money(c_bono, bono_base)
    c_bono.font = Font(bold=True)
    c_bono.border = border
    ws.cell(row=r, column=3, value="×0,5 si 1/1 cumplido · ×1 si todos cumplidos (≥2)").border = border
    r += 2

    # ── Detalle liquidación ────────────────────────────────────────────────────
    section_bar("DETALLE DE LIQUIDACIÓN POR OBJETIVO")
    headers_detalle = [
        "ID Objetivo",
        "ID Vendedor",
        "Vendedor",
        "Tipo",
        "Fecha límite",
        "Descripción",
        "Meta",
        "Avance",
        "PDVs dist.",
        "% Cumpl.",
        "Resultado",
        "Tarifa unit.",
        "Monto liquidado",
    ]
    _xlsx_write_header_row(ws, headers_detalle, r, mods, ncols=NCOLS)
    detalle_header_row = r
    r += 1

    vendedor_grupos: dict[int, list[dict]] = OrderedDict()
    for row in liq["vendedores"]:
        vid = row["id_vendedor"]
        vendedor_grupos.setdefault(vid, []).append(row)

    data_idx = 0
    for vid, objs in vendedor_grupos.items():
        nombre_vend = objs[0].get("nombre_vendedor") or f"Vendedor {vid}"
        subtotal_monto = 0.0

        for obj_row in objs:
            zebra = fill_zebra if data_idx % 2 == 0 else None
            data_idx += 1
            avance_pdvs = obj_row.get("avance_pdvs")
            vals = [
                obj_row.get("id_objetivo", ""),
                vid,
                nombre_vend,
                obj_row.get("tipo_label") or _tipo_label(str(obj_row.get("tipo") or "")),
                obj_row.get("fecha_objetivo") or "",
                obj_row.get("descripcion") or "",
                obj_row.get("meta", 0),
                obj_row.get("avance", 0),
                avance_pdvs if avance_pdvs is not None else "—",
                obj_row.get("pct", 0),
                str(obj_row.get("resultado_final") or "exito").upper(),
                obj_row.get("tarifa_unitaria", 0),
                obj_row.get("monto", 0),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=(col == 6))
                if zebra:
                    cell.fill = zebra
                if col == 10:
                    _xlsx_pct(cell, val if isinstance(val, (int, float)) else 0)
                elif col in (12, 13):
                    _xlsx_money(cell, val if isinstance(val, (int, float)) else 0)
            subtotal_monto += _safe_float(obj_row.get("monto"))
            r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        sub_c = ws.cell(row=r, column=1, value=f"SUBTOTAL — {nombre_vend} ({len(objs)} objetivo(s))")
        sub_c.font = Font(bold=True, color="5B21B6")
        sub_c.fill = fill_subtotal
        sub_c.border = border
        for col in range(2, 12):
            ws.cell(row=r, column=col).fill = fill_subtotal
            ws.cell(row=r, column=col).border = border
        sub_m = ws.cell(row=r, column=NCOLS)
        _xlsx_money(sub_m, subtotal_monto)
        sub_m.font = Font(bold=True, color="5B21B6")
        sub_m.fill = fill_subtotal
        sub_m.border = border
        r += 1

    r += 1

    # ── Mando medio ────────────────────────────────────────────────────────────
    section_bar("BONO MANDO MEDIO")
    headers_mando = [
        "ID Vendedor",
        "Vendedor",
        "Objetivos asignados",
        "Objetivos cumplidos",
        "Factor bono",
        "Bono base",
        "Monto bono",
    ]
    _xlsx_write_header_row(ws, headers_mando, r, mods, ncols=7)
    mando_header_row = r
    r += 1

    for mm_row in liq["mando_medio"]:
        factor = _safe_float(mm_row.get("factor"))
        factor_txt = "×0,5 (1/1)" if factor == 0.5 else ("×1 (todos)" if factor == 1.0 else "—")
        nombre_mm = mm_row.get("nombre_vendedor") or f"Vendedor {mm_row['id_vendedor']}"
        row_vals = [
            mm_row["id_vendedor"],
            nombre_mm,
            mm_row.get("asignados", 0),
            mm_row.get("cumplidos", 0),
            factor_txt,
            bono_base,
            mm_row.get("monto_bono", 0),
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            if col in (6, 7):
                _xlsx_money(cell, val if isinstance(val, (int, float)) else 0)
            if factor <= 0 and col == 7:
                cell.font = Font(color="9CA3AF")
        r += 1

    r += 1

    # ── Totales ──────────────────────────────────────────────────────────────
    section_bar("TOTALES")
    totales = [
        ("Total pagos por objetivos cumplidos", liq["total_vendedores"]),
        ("Total bono mando medio", liq["total_mando_medio"]),
        ("TOTAL A LIQUIDAR — DISTRIBUIDORA", liq["total_distribuidora"]),
    ]
    for i, (label, monto) in enumerate(totales):
        is_grand = i == len(totales) - 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS - 1)
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = font_total if is_grand else Font(bold=True, size=11)
        lc.fill = fill_total
        lc.border = border
        for col in range(2, NCOLS):
            ws.cell(row=r, column=col).fill = fill_total
            ws.cell(row=r, column=col).border = border
        mc = ws.cell(row=r, column=NCOLS)
        _xlsx_money(mc, monto)
        mc.font = font_total if is_grand else Font(bold=True, size=11)
        mc.fill = fill_total
        mc.border = border
        ws.row_dimensions[r].height = 22 if is_grand else 18
        r += 1

    ws.freeze_panes = ws.cell(row=detalle_header_row + 1, column=1)
    _xlsx_autofit(ws, max_width=48)
    ws.column_dimensions[get_column_letter(6)].width = 36  # descripción

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Job de archivado ─────────────────────────────────────────────────────────

def archivar_terminados_compania_7d(dist_id: int | None = None) -> dict:
    """
    Job de limpieza: marca liquidacion_at en objetivos compañía cumplidos
    hace más de 7 días y aún sin liquidacion_at.

    Si dist_id es None, recorre todos los distribuidores.
    """
    now_utc = datetime.now(timezone.utc)
    cutoff = now_utc - timedelta(days=7)
    cutoff_iso = cutoff.isoformat()
    now_iso = now_utc.isoformat()

    # Obtener distribuidores a procesar
    if dist_id is not None:
        dist_ids = [dist_id]
    else:
        try:
            rows = (
                sb.table("distribuidores")
                .select("id_distribuidor")
                .execute()
                .data
                or []
            )
            dist_ids = [_safe_int(r.get("id_distribuidor")) for r in rows if r.get("id_distribuidor")]
        except Exception:
            # Fallback: columna "id"
            try:
                rows = (
                    sb.table("distribuidores")
                    .select("id")
                    .execute()
                    .data
                    or []
                )
                dist_ids = [_safe_int(r.get("id")) for r in rows if r.get("id")]
            except Exception as e:
                logger.error("[archivar_terminados] No se pudieron obtener distribuidores: %s", e)
                return {"archivados": 0, "errores": 1}

    total_archivados = 0
    total_errores = 0

    for d_id in dist_ids:
        try:
            # Buscar objetivos elegibles para este distribuidor
            def _query_terminados(offset: int):
                return (
                    sb.table(_OBJETIVOS_TABLE)
                    .select("id")
                    .eq("id_distribuidor", d_id)
                    .eq("origen", "compania")
                    .eq("cumplido", True)
                    .lt("completed_at", cutoff_iso)
                    .is_("liquidacion_at", "null")
                    .range(offset, offset + PAGE - 1)
                    .execute()
                    .data
                    or []
                )

            obj_rows = _paginate(_query_terminados)
            obj_ids = [str(r["id"]) for r in obj_rows if r.get("id")]

            for obj_id in obj_ids:
                try:
                    sb.table(_OBJETIVOS_TABLE).update(
                        {"liquidacion_at": now_iso}
                    ).eq("id", obj_id).execute()
                    total_archivados += 1
                except Exception as e_upd:
                    logger.warning(
                        "[archivar_terminados] Error archivando obj=%s dist=%s: %s",
                        obj_id,
                        d_id,
                        e_upd,
                    )
                    total_errores += 1

        except Exception as e_dist:
            logger.warning(
                "[archivar_terminados] Error procesando dist=%s: %s", d_id, e_dist
            )
            total_errores += 1

    logger.info(
        "[archivar_terminados] archivados=%s errores=%s dists=%s",
        total_archivados,
        total_errores,
        len(dist_ids),
    )
    return {"archivados": total_archivados, "errores": total_errores}
