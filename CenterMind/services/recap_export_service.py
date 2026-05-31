# -*- coding: utf-8 -*-
"""
Export XLSX del Repaso Comercial.

Genera un XLSX con 2 hojas:
  1. Resumen vendedores
  2. Detalle operativo (altas + artículos top apilados)
"""
from __future__ import annotations

import io
import logging
from typing import Any

from services.recap_snapshot_service import list_recaps_for_mes

logger = logging.getLogger("recap_export_service")

_RANKING_PERIOD_PRIORITY = ("C", "Q2", "Q1")


def _ranking_period_label(mes: str, period_suffix: str) -> str:
    if period_suffix == "C":
        return f"Cierre de mes {mes}"
    if period_suffix == "Q2":
        return f"2da quincena {mes} (provisional)"
    return f"1ra quincena {mes} (provisional)"


def _pick_ranking_snapshots(snapshots: list[dict], mes: str) -> tuple[list[dict], str]:
    """Elige snapshots del cierre (C) o el período más reciente disponible en el mes."""
    by_period: dict[str, list[dict]] = {}
    for snap in snapshots:
        pk = str(snap.get("periodo_key") or "")
        if not pk.startswith(f"{mes}-"):
            continue
        suffix = pk.rsplit("-", 1)[-1]
        by_period.setdefault(suffix, []).append(snap)

    for suffix in _RANKING_PERIOD_PRIORITY:
        if by_period.get(suffix):
            return by_period[suffix], _ranking_period_label(mes, suffix)

    return [], f"Mes {mes}"


def _build_ranking_oficial_rows(snapshots: list[dict], mes: str) -> tuple[list[dict], str]:
    """
    Ranking oficial del mes según score FIFA de los snapshots de repaso.
    Prioriza cierre (C); si no existe, usa Q2 o Q1.
    """
    chosen, label = _pick_ranking_snapshots(snapshots, mes)
    rows: list[dict] = []
    for snap in chosen:
        payload = snap.get("payload") or {}
        if not isinstance(payload, dict):
            continue
        carta = payload.get("carta") or {}
        raw = carta.get("raw_kpis") or {}
        rows.append(
            {
                "nombre": (carta.get("nombre") or "").strip(),
                "sucursal": (carta.get("sucursal") or "").strip(),
                "score": _safe_float(carta.get("score")),
                "pdvs": _safe_int(raw.get("pdvs")),
                "altas": _safe_int(raw.get("altas")),
                "exhibiciones": _safe_int(raw.get("exhibiciones")),
                "compradores": _safe_int(raw.get("compradores")),
                "bultos": _safe_float(raw.get("bultos")),
                "cobertura_pct": _safe_float(raw.get("cobertura_pct")),
            }
        )

    rows.sort(key=lambda r: (-r["score"], r["nombre"].lower()))
    for i, row in enumerate(rows, 1):
        row["puesto"] = i
    return rows, label


def _try_build_ranking_live(dist_id: int, mes: str) -> list[dict] | None:
    """Fallback: ranking en vivo de estadísticas (mismo orden que el portal)."""
    try:
        from services.estadisticas_service import build_carta_resumen

        cards = build_carta_resumen(dist_id, [mes], None)
    except Exception as e:
        logger.warning("[recap_export] ranking live dist=%s mes=%s: %s", dist_id, mes, e)
        return None

    rows: list[dict] = []
    for i, card in enumerate(cards, 1):
        raw = card.get("raw_kpis") or {}
        rows.append(
            {
                "puesto": i,
                "nombre": (card.get("nombre") or "").strip(),
                "sucursal": (card.get("sucursal") or "").strip(),
                "score": round(_safe_float(card.get("score")), 2),
                "pdvs": _safe_int(raw.get("pdvs")),
                "altas": _safe_int(raw.get("altas")),
                "exhibiciones": _safe_int(raw.get("exhibiciones")),
                "compradores": _safe_int(raw.get("compradores")),
                "bultos": _safe_float(raw.get("bultos_raw") or raw.get("bultos")),
                "cobertura_pct": _safe_float(raw.get("cobertura_pct")),
            }
        )
    return rows



def _wb():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        return openpyxl, Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl no instalado — no se puede generar XLSX")


def _header_row(ws, row: list[str], openpyxl_mods, row_num: int = 1):
    _, Font, PatternFill, Alignment = openpyxl_mods
    violet = "7C3AED"
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(fill_type="solid", fgColor=violet)
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws, max_width: int = 50, start_row: int = 1):
    for col in ws.columns:
        cells = [c for c in col if c.row >= start_row]
        if not cells:
            continue
        max_len = max((len(str(c.value or "")) for c in cells), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, max_width)


def _safe_float(val: Any, fallback: float = 0.0) -> float:
    try:
        return float(val)
    except (TypeError, ValueError):
        return fallback


def _safe_int(val: Any, fallback: int = 0) -> int:
    try:
        return int(val)
    except (TypeError, ValueError):
        return fallback


def export_recap_dist_mes_xlsx(dist_id: int, mes: str) -> bytes:
    """
    Genera XLSX con:
      1. Ranking oficial del mes (score FIFA, orden portal)
      2. Resumen vendedores por período
      3. Detalle operativo (altas + artículos top)
    """
    snapshots = list_recaps_for_mes(dist_id, mes)
    mods = _wb()
    openpyxl = mods[0]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore

    # ── Hoja 1: Ranking oficial del mes ────────────────────────────────────────
    ranking_rows, ranking_label = _build_ranking_oficial_rows(snapshots, mes)
    ranking_source = ranking_label
    if not ranking_rows:
        live = _try_build_ranking_live(dist_id, mes)
        if live:
            ranking_rows = live
            ranking_source = f"Estadísticas en vivo — mes {mes}"

    ws_rank = wb.create_sheet("Ranking oficial mes", 0)
    ws_rank.cell(row=1, column=1, value=f"Ranking oficial — {ranking_source}")
    ws_rank.cell(row=1, column=1).font = mods[1](bold=True, size=11)
    headers_rank = [
        "Puesto",
        "Vendedor",
        "Sucursal",
        "Score FIFA",
        "PDVs",
        "Altas",
        "Exhibiciones",
        "Compradores",
        "Bultos",
        "Cobertura %",
    ]
    _header_row(ws_rank, headers_rank, mods, row_num=3)

    for i, row in enumerate(ranking_rows, 4):
        ws_rank.append([
            row["puesto"],
            row["nombre"],
            row["sucursal"],
            round(row["score"], 2),
            row["pdvs"],
            row["altas"],
            row["exhibiciones"],
            row["compradores"],
            round(row["bultos"], 2),
            round(row["cobertura_pct"], 1),
        ])

    # ── Hoja 2: Resumen vendedores ─────────────────────────────────────────────
    ws1 = wb.create_sheet("Resumen vendedores")
    headers1 = [
        "Vendedor",
        "Sucursal",
        "Período",
        "Score",
        "PDVs",
        "Altas",
        "Exhibiciones Enviadas",
        "Aprobadas",
        "Destacadas",
        "Compradores",
        "Bultos Total",
        "Δ Score vs Ant",
        "Δ Exhibiciones vs Ant",
        "Flag calidad ERP",
    ]
    _header_row(ws1, headers1, mods)

    # ── Hoja 2: Detalle operativo ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle operativo")
    headers2 = ["Sección", "Vendedor", "Período", "Campo", "Valor"]
    _header_row(ws2, headers2, mods)

    detail_row = 2  # empezamos desde la fila 2 (1 = header)

    for snap in snapshots:
        payload: dict = snap.get("payload") or {}
        periodo_key: str = snap.get("periodo_key") or payload.get("periodo_key") or ""
        carta: dict = payload.get("carta") or {}
        carta_anterior: dict | None = payload.get("carta_anterior")
        raw = carta.get("raw_kpis") or {}
        raw_ant = (carta_anterior or {}).get("raw_kpis") or {}

        # Nombre y sucursal
        nombre = carta.get("nombre") or ""
        sucursal = carta.get("sucursal") or ""

        # Tipo período
        tipo_periodo = periodo_key.rsplit("-", 1)[-1] if periodo_key else ""

        # Score
        score = _safe_float(carta.get("score"))
        score_ant = _safe_float(carta_anterior.get("score") if carta_anterior else None)
        delta_score: str = (
            f"{score - score_ant:+.2f}" if carta_anterior is not None else "N/A"
        )

        # KPIs
        pdvs = _safe_int(raw.get("pdvs"))
        altas_count = _safe_int(raw.get("altas"))
        ex_data: dict = payload.get("exhibiciones") or {}
        ex_total = _safe_int(ex_data.get("total_logicas"))
        ex_aprobadas = _safe_int(ex_data.get("aprobadas"))
        ex_destacadas = _safe_int(ex_data.get("destacadas"))
        compradores = _safe_int(raw.get("compradores"))
        bultos_total = _safe_float(payload.get("bultos_total"))

        # Delta exhibiciones vs anterior
        ex_ant = _safe_int((payload.get("carta_anterior") or {}).get("raw_kpis", {}).get("exhibiciones") if payload.get("carta_anterior") else None)
        delta_ex: str = (
            f"{ex_total - ex_ant:+d}" if payload.get("carta_anterior") is not None else "N/A"
        )

        # Flag calidad ERP
        dq: dict = payload.get("data_quality") or {}
        erp_flag = "OK" if dq.get("erp_sync_ok", True) else "ALERTA"

        ws1.append([
            nombre,
            sucursal,
            tipo_periodo,
            round(score, 2),
            pdvs,
            altas_count,
            ex_total,
            ex_aprobadas,
            ex_destacadas,
            compradores,
            round(bultos_total, 2),
            delta_score,
            delta_ex,
            erp_flag,
        ])

        # ── Hoja 2: Altas ──────────────────────────────────────────────────────
        altas: list[dict] = payload.get("altas") or []
        for alta in altas:
            ws2.cell(row=detail_row, column=1, value="Altas")
            ws2.cell(row=detail_row, column=2, value=nombre)
            ws2.cell(row=detail_row, column=3, value=tipo_periodo)
            ws2.cell(row=detail_row, column=4, value=alta.get("nombre") or alta.get("id_cliente_erp") or "")
            ws2.cell(row=detail_row, column=5, value=alta.get("fecha_alta") or "")
            detail_row += 1

        # ── Hoja 2: Artículos Top ──────────────────────────────────────────────
        bultos_top: list[dict] = payload.get("bultos_top") or []
        for bt in bultos_top:
            ws2.cell(row=detail_row, column=1, value="Artículos Top")
            ws2.cell(row=detail_row, column=2, value=nombre)
            ws2.cell(row=detail_row, column=3, value=tipo_periodo)
            ws2.cell(row=detail_row, column=4, value=bt.get("cod_articulo") or bt.get("articulo") or "")
            ws2.cell(row=detail_row, column=5, value=bt.get("bultos") or 0)
            detail_row += 1

    _autofit(ws_rank, start_row=3)
    _autofit(ws1)
    _autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
