# -*- coding: utf-8 -*-
"""
Export service para snapshots de reportería.
Genera XLSX por source (sigo / comprobantes / bultos) desde el snapshot en memoria.
"""
from __future__ import annotations

import io
import logging
from typing import Any

logger = logging.getLogger("ShelfyAPI")


def _wb():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        return openpyxl, Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl no instalado — no se puede generar XLSX")


def _header_row(ws, row: list[str], openpyxl_mods):
    _, Font, PatternFill, Alignment = openpyxl_mods
    violet = "7C3AED"
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=1, column=col_idx, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(fill_type="solid", fgColor=violet)
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws, max_width: int = 50):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, max_width)


def export_xlsx(snapshot: dict[str, Any]) -> bytes:
    """Genera XLSX desde el snapshot. Soporta sigo, comprobantes, bultos."""
    source = snapshot.get("source", "")
    mods = _wb()
    openpyxl = mods[0]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore

    # ── Hoja KPIs ─────────────────────────────────────────────────────────────
    ws_kpi = wb.create_sheet("KPIs")
    _header_row(ws_kpi, ["Indicador", "Valor", "Unidad"], mods)
    for kpi in snapshot.get("kpis") or []:
        ws_kpi.append([kpi.get("label"), kpi.get("value"), kpi.get("unit", "")])
    _autofit(ws_kpi)

    # ── Sheets por source ─────────────────────────────────────────────────────
    if source == "sigo":
        _export_sigo(wb, snapshot, mods)
    elif source == "comprobantes":
        _export_comprobantes(wb, snapshot, mods)
    elif source == "comprobantes_detallado":
        _export_comprobantes_detallado(wb, snapshot, mods)
    elif source == "bultos":
        _export_bultos(wb, snapshot, mods)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _export_sigo(wb, snapshot: dict, mods):
    # Hoja por_vendedor_y_dia
    ws = wb.create_sheet("Vendedor × Día")
    cols = ["Vendedor", "Fecha", "Planeadas", "Ejecutadas", "Sin Visita",
            "Con Venta", "Motivo No Venta", "Sin Info",
            "H. Primera Visita", "H. Primera Venta", "T. Prom. Venta (min)"]
    _header_row(ws, cols, mods)
    for r in snapshot.get("por_vendedor_y_dia") or []:
        ws.append([
            r.get("vendedor"), r.get("fecha"),
            r.get("planeadas"), r.get("ejecutadas"), r.get("sin_visita"),
            r.get("con_venta"), r.get("motivo_no_venta"), r.get("sin_info"),
            r.get("hora_primera_visita"), r.get("hora_primera_venta"),
            r.get("tiempo_promedio_venta_min"),
        ])
    _autofit(ws)

    # Hoja ranking vendedores
    ws2 = wb.create_sheet("Ranking Vendedores")
    _header_row(ws2, ["Vendedor", "Cobertura %", "Visitados / Total", "Efectividad %", "Ventas"], mods)
    for r in snapshot.get("top_clientes") or []:
        ws2.append([
            r.get("nombre_cliente"),
            r.get("importe_total"),
            r.get("vendedor_nombre"),
            r.get("sucursal_nombre"),
            r.get("cantidad_facturas"),
        ])
    _autofit(ws2)

    # Hoja por sucursal (if available)
    if snapshot.get("por_sucursal"):
        ws3 = wb.create_sheet("Por Sucursal")
        _header_row(ws3, ["Sucursal", "Total", "Visitados", "Ventas", "Cobertura %", "Efectividad %"], mods)
        for r in snapshot.get("por_sucursal") or []:
            ws3.append([r.get("sucursal"), r.get("total"), r.get("visitados"),
                        r.get("ventas"), r.get("cobertura"), r.get("efectividad")])
        _autofit(ws3)

    # Hoja por hora (if available)
    if snapshot.get("por_hora"):
        ws4 = wb.create_sheet("Por Hora")
        _header_row(ws4, ["Hora", "Visitas", "Ventas"], mods)
        for r in snapshot.get("por_hora") or []:
            ws4.append([r.get("hora"), r.get("visitas"), r.get("ventas")])
        _autofit(ws4)


def _export_comprobantes(wb, snapshot: dict, mods):
    # Hoja top clientes
    ws = wb.create_sheet("Top Clientes")
    _header_row(ws, ["Cliente", "Vendedor", "Sucursal", "Facturas", "Importe Total"], mods)
    for r in snapshot.get("top_clientes") or []:
        ws.append([
            r.get("nombre_cliente"), r.get("vendedor_nombre"), r.get("sucursal_nombre"),
            r.get("cantidad_facturas"), r.get("importe_total"),
        ])
    _autofit(ws)

    # Hoja ranking vendedores
    ws2 = wb.create_sheet("Ranking Vendedores")
    _header_row(ws2, ["Vendedor", "Importe Total"], mods)
    for r in snapshot.get("top_vendedores") or []:
        ws2.append([r.get("nombre"), r.get("valor")])
    _autofit(ws2)

    # Hoja clientes full (if available)
    if snapshot.get("clientes_full"):
        ws3 = wb.create_sheet("Clientes Full")
        _header_row(ws3, ["Cliente", "Vendedor", "Sucursal", "Canal", "Importe", "Contado", "Cta Cte", "Operaciones"], mods)
        for r in snapshot.get("clientes_full") or []:
            ws3.append([r.get("nombre_cliente"), r.get("vendedor"), r.get("sucursal"),
                        r.get("canal"), r.get("importe"), r.get("contado"),
                        r.get("cc"), r.get("n_ops")])
        _autofit(ws3)

    # Hoja por canal (if available)
    if snapshot.get("por_canal"):
        ws4 = wb.create_sheet("Por Canal")
        _header_row(ws4, ["Canal", "Subcanal", "Importe", "Contado", "Cta Cte", "Operaciones"], mods)
        for r in snapshot.get("por_canal") or []:
            ws4.append([r.get("canal"), r.get("subcanal"), r.get("importe"),
                        r.get("contado"), r.get("cc"), r.get("n_ops")])
        _autofit(ws4)


def _export_comprobantes_detallado(wb, snapshot: dict, mods):
    # Hoja artículos
    ws = wb.create_sheet("Artículos")
    _header_row(ws, ["Artículo", "Importe", "Operaciones", "Clientes", "Prom/Sem"], mods)
    for r in snapshot.get("por_articulo") or []:
        ws.append([r.get("articulo"), r.get("importe"), r.get("n_ops"),
                   r.get("n_clientes"), r.get("prom_sem")])
    _autofit(ws)

    # Hoja vendedores × artículo
    ws2 = wb.create_sheet("Vendedores × Artículo")
    _header_row(ws2, ["Vendedor", "Artículo", "Importe", "Operaciones"], mods)
    for r in snapshot.get("por_vendedor_articulo") or []:
        ws2.append([r.get("vendedor"), r.get("articulo"), r.get("importe"), r.get("n_ops")])
    _autofit(ws2)

    # Hoja clientes × artículo
    ws3 = wb.create_sheet("Clientes × Artículo")
    _header_row(ws3, ["Cliente", "Artículo", "Importe", "Operaciones"], mods)
    for r in snapshot.get("clientes_x_articulo") or []:
        ws3.append([r.get("cliente"), r.get("articulo"), r.get("importe"), r.get("n_ops")])
    _autofit(ws3)

    # Hoja top clientes
    ws4 = wb.create_sheet("Top Clientes")
    _header_row(ws4, ["Cliente", "Vendedor", "Sucursal", "Importe", "Operaciones"], mods)
    for r in snapshot.get("top_clientes") or []:
        ws4.append([r.get("nombre_cliente"), r.get("vendedor_nombre"),
                    r.get("sucursal_nombre"), r.get("importe_total"), r.get("cantidad_facturas")])
    _autofit(ws4)


def _export_bultos(wb, snapshot: dict, mods):
    # Hoja top PDVs
    ws = wb.create_sheet("Top PDVs")
    _header_row(ws, ["PDV / Cliente", "Vendedor", "Sucursal", "Bultos Totales", "Prom/Sem"], mods)
    for r in snapshot.get("top_clientes") or []:
        ws.append([
            r.get("nombre_cliente"), r.get("vendedor_nombre"), r.get("sucursal_nombre"),
            r.get("cantidad_facturas"), r.get("importe_total"),
        ])
    _autofit(ws)

    # Hoja top artículos
    ws2 = wb.create_sheet("Top Artículos")
    _header_row(ws2, ["Artículo", "Bultos Totales"], mods)
    for r in snapshot.get("top_vendedores") or []:
        ws2.append([r.get("nombre"), r.get("valor")])
    _autofit(ws2)

    # Hoja por vendedor bultos (if available)
    if snapshot.get("por_vendedor_bultos"):
        ws3 = wb.create_sheet("Por Vendedor")
        _header_row(ws3, ["Vendedor", "Bultos", "Prom/Sem", "Clientes", "% >2.5/sem"], mods)
        for r in snapshot.get("por_vendedor_bultos") or []:
            ws3.append([r.get("vendedor"), r.get("bultos"), r.get("prom_sem"),
                        r.get("n_clientes"), r.get("pct_25")])
        _autofit(ws3)
